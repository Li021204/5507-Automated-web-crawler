import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import openpyxl
import openpyxl.styles
from typing import List, Optional, Tuple
import json
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import quote
import datetime as dt

import douban


def crawl_zol_products(start_page=1, end_page=3):
    """
    爬取中关村在线手机产品信息
    
    参数:
        start_page: 起始页码，默认为1
        end_page: 结束页码，默认为3
    
    返回:
        products: 产品信息列表
    """
    # 基础URL模板
    base_url = "https://detail.zol.com.cn/cell_phone_index/subcate57_0_list_1_0_1_2_0_{}.html"
    
    products = []
    
    # 爬取多页数据
    for page in range(start_page, end_page + 1):
        url = base_url.format(page)
        
        try:
            # 发送请求
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
            resp = requests.get(url, headers=headers, timeout=10)
            resp.encoding = 'gbk'
            
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, 'html.parser')
                items3 = soup.select("li[data-follow-id]")
                
                print(f"正在爬取第{page}页，找到{len(items3)}个产品")
                
                for phone in items3:
                    try:
                        # 获取完整的title
                        full_title = phone.h3.a["title"].strip()
                        
                        # 拆分手机型号和产品特点
                        # 方法1: 从HTML结构中直接提取
                        model_elem = phone.h3.a.contents[0] if phone.h3.a.contents else ""
                        model = model_elem.strip() if isinstance(model_elem, str) else ""
                        
                        features_elem = phone.h3.a.find('span')
                        features = features_elem.text.strip() if features_elem else ""
                        
                        # 方法2: 如果方法1不成功，尝试从完整title中拆分
                        if not model or not features:
                            # 假设title中型号和特点之间有明显的分隔
                            if " " in full_title:
                                parts = full_title.split(" ", 1)
                                model = parts[0].strip() if not model else model
                                features = parts[1].strip() if not features else features
                            else:
                                model = full_title
                        
                        link = phone.h3.a["href"].strip()
        
                        # 修正链接
                        if link.startswith("/"):
                            link = "https://detail.zol.com.cn" + link
                        elif link.startswith("//"):
                            link = "https:" + link
        
                        # 提取价格（包含符号）
                        price_elem = phone.select_one('.price-normal')
                        if price_elem:
                            price_sign = price_elem.select_one('.price-sign')
                            price_type = price_elem.select_one('.price-type')
                            price = f"{price_sign.text if price_sign else '¥'}{price_type.text if price_type else ''}"
                        else:
                            price = "暂无价格"
        
                        # 提取评分
                        score_elem = phone.select_one('.score')
                        score = score_elem.text.strip() if score_elem else "暂无评分"
        
                        products.append({
                            "model": model,  # 手机型号
                            "features": features,  # 产品特点
                            "price": price,
                            "score": score,
                            "link": link
                        })
                        
                    except Exception as e:
                        print(f"处理单个产品时出错: {e}")
                        continue
                
                print(f"第{page}页爬取完成")
                
                # 添加延迟，避免请求过快
                time.sleep(1)
                
            else:
                print(f"第{page}页请求失败，状态码: {resp.status_code}")
                
        except Exception as e:
            print(f"爬取第{page}页时出错: {e}")
            continue
    
    print(f"总共爬取到 {len(products)} 个产品")
    return products


def export_to_excel(df, filename=None, formatted=True):
    """
    导出数据到Excel文件
    
    参数:
        df: pandas DataFrame对象
        filename: 文件名，如果为None则自动生成带时间戳的文件名
        formatted: 是否使用格式化导出（包含列宽设置等）
    
    返回:
        filename: 生成的文件名
    """
    try:
        if filename is None:
            # 生成带时间戳的文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"中关村在线产品信息_{timestamp}.xlsx"
        
        if formatted:
            # 使用格式化导出
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(
                    writer,
                    index=False,
                    sheet_name='产品信息'
                )
                
                # 获取工作表对象
                worksheet = writer.sheets['产品信息']
                
                # 设置列宽
                worksheet.column_dimensions['A'].width = 40  # model列
                worksheet.column_dimensions['B'].width = 80  # features列  
                worksheet.column_dimensions['C'].width = 15  # price列
                worksheet.column_dimensions['D'].width = 15  # score列
                worksheet.column_dimensions['E'].width = 70  # link列
                
                # 设置标题行格式（可选）
                header_fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                for cell in worksheet[1]:
                    cell.fill = header_fill
            
            print(f"数据已成功导出为格式化Excel文件：{filename}")
        else:
            # 简单导出
            df.to_excel(
                filename,
                index=False,           # 不包含行索引
                sheet_name='产品信息'   # 工作表名称
            )
            print(f"数据已成功导出为Excel文件：{filename}")
        
        print(f"共导出 {len(df)} 个产品")
        return filename
        
    except Exception as e:
        print(f"导出Excel文件时出错: {e}")
        raise


def read_phone_models_from_excel(excel_file_path: str) -> List[str]:
    """
    从Excel文件中读取phone model列的内容
    
    参数:
        excel_file_path: Excel文件路径
    
    返回:
        phone_models: 手机型号列表
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(excel_file_path, sheet_name='产品信息', engine='openpyxl')
        
        # 检查列名，支持 'model' 或 'phone model'
        if 'model' in df.columns:
            column_name = 'model'
        elif 'phone model' in df.columns:
            column_name = 'phone model'
        else:
            # 尝试查找包含 'model' 的列
            model_columns = [col for col in df.columns if 'model' in col.lower()]
            if model_columns:
                column_name = model_columns[0]
                print(f"找到列名: {column_name}")
            else:
                raise ValueError(f"Excel文件中未找到 'model' 或 'phone model' 列。可用列: {list(df.columns)}")
        
        # 提取手机型号，去除空值和重复值
        phone_models = df[column_name].dropna().astype(str).str.strip()
        phone_models = phone_models[phone_models != ''].unique().tolist()
        
        print(f"从Excel文件中读取到 {len(phone_models)} 个手机型号")
        return phone_models
        
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        raise


def run_douban_workflow(keywords: List[str], headless: bool = False) -> int:
    """
    依次爬取豆瓣小组搜索结果。
    返回成功采集的关键词数量。
    """
    if not keywords:
        print("[Douban] 关键词列表为空，跳过豆瓣爬取。")
        return 0

    # 预清洗：去掉括号内容后再搜索
    keywords = [douban_sanitize_keyword(k) for k in keywords]
    keywords = [k for k in keywords if k]

    os.makedirs(DOUBAN_OUTPUT_DIR, exist_ok=True)
    output_dir = os.path.abspath(DOUBAN_OUTPUT_DIR)

    original_headless = DOUBAN_HEADLESS
    if headless is not None:
        globals()["DOUBAN_HEADLESS"] = headless

    try:
        driver = douban_build_driver()
    except Exception as e:
        print(f"[Douban] 初始化浏览器失败: {e}")
        return 0

    try:
        douban_inject_manual_cookie_if_any(driver)

        if not douban_ensure_logged_in(driver, DOUBAN_QR_LOGIN_WAIT_SECONDS):
            print("[Douban] 登录校验失败，豆瓣爬虫将跳过。")
            return 0

        success_count = 0
        total = len(keywords)
        all_data: List[tuple] = []

        for idx, keyword in enumerate(keywords, 1):
            print(f"\n[Douban] 进度：{idx}/{total} | 关键词：{keyword}")
            data = douban_crawl_keyword(driver, keyword, output_dir)
            if data:
                success_count += 1
                all_data.append((keyword, data))
                douban_save_single_excel(keyword, data, output_dir)
            if idx < total:
                pause = douban_rand_in_range((2.0, 4.0))
                print(f"[Douban] 等待 {pause:.1f} 秒后处理下一个关键词...")
                douban_human_pause(2.0, 4.0)

        if all_data:
            douban_save_combined_excel(all_data, output_dir)

        return success_count
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        globals()["DOUBAN_HEADLESS"] = original_headless


def step3_merge_all_excel(
    run_xhs: bool = True,
    run_weibo: bool = True,
    run_douban: bool = True,
    run_zhihu: bool = True,
    run_bilibili: bool = True,
) -> Optional[str]:
    """
    Step 3: 合并所有平台的整合Excel文件到一个汇总Excel中，每个平台一个sheet。
    
    参数:
        run_xhs: 是否包含小红书数据
        run_weibo: 是否包含微博数据
        run_douban: 是否包含豆瓣数据
        run_zhihu: 是否包含知乎数据
        run_bilibili: 是否包含B站数据
    
    返回:
        汇总Excel文件路径，如果失败则返回None
    """
    import glob
    
    try:
        # 查找各平台最新的整合Excel文件
        platform_files = {}
        
        # 1. 小红书
        if run_xhs:
            xhs_pattern = os.path.join(XHS_OUTPUT_DIR, "xhs_整合_*.xlsx")
            xhs_files = glob.glob(xhs_pattern)
            if xhs_files:
                xhs_files.sort(key=os.path.getmtime, reverse=True)
                platform_files["小红书"] = xhs_files[0]
        
        # 2. 微博
        if run_weibo:
            weibo_pattern = os.path.join(WEIBO_OUTPUT_DIR, "微博搜索结果_整合_*.xlsx")
            weibo_files = glob.glob(weibo_pattern)
            if weibo_files:
                weibo_files.sort(key=os.path.getmtime, reverse=True)
                platform_files["微博"] = weibo_files[0]
        
        # 3. 豆瓣
        if run_douban:
            douban_pattern = os.path.join(DOUBAN_OUTPUT_DIR, "豆瓣搜索结果_整合_*.xlsx")
            douban_files = glob.glob(douban_pattern)
            if douban_files:
                douban_files.sort(key=os.path.getmtime, reverse=True)
                platform_files["豆瓣"] = douban_files[0]
        
        # 4. 知乎
        if run_zhihu:
            zhihu_file = os.path.join(ZHIHU_OUTPUT_DIR, "zhihu_all_combined.xlsx")
            if os.path.exists(zhihu_file):
                platform_files["知乎"] = zhihu_file
        
        # 5. B站
        if run_bilibili:
            bili_file = os.path.join(BILIBILI_OUTPUT_DIR, "bilibili_all.xlsx")
            if os.path.exists(bili_file):
                platform_files["B站"] = bili_file
        
        # 6. 中关村（Step1的输出）
        zol_pattern = "中关村在线产品信息_*.xlsx"
        zol_files = glob.glob(zol_pattern)
        if zol_files:
            zol_files.sort(key=os.path.getmtime, reverse=True)
            platform_files["中关村"] = zol_files[0]
        
        if not platform_files:
            print("[Step3] 未找到任何平台的整合Excel文件，跳过合并。")
            return None
        
        # 创建汇总Excel文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_file = os.path.join(DEFAULT_OUTPUT_DIR, f"所有平台汇总_{timestamp}.xlsx")
        os.makedirs(DEFAULT_OUTPUT_DIR, exist_ok=True)
        
        print(f"[Step3] 找到 {len(platform_files)} 个平台的整合文件，开始合并...")
        for platform, filepath in platform_files.items():
            print(f"  - {platform}: {filepath}")
        
        with pd.ExcelWriter(merged_file, engine="openpyxl") as writer:
            for platform, filepath in platform_files.items():
                try:
                    # 读取Excel文件
                    if platform == "B站":
                        # B站是多表结构，读取所有sheet
                        excel_file = pd.ExcelFile(filepath)
                        for sheet_name in excel_file.sheet_names:
                            df = pd.read_excel(filepath, sheet_name=sheet_name)
                            if not df.empty:
                                safe_sheet_name = f"B站_{sheet_name}"[:31]  # Excel sheet名称限制31字符
                                df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                    else:
                        # 其他平台是单表结构
                        df = pd.read_excel(filepath)
                        if not df.empty:
                            safe_sheet_name = platform[:31]  # Excel sheet名称限制31字符
                            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                            print(f"[Step3] ✓ {platform}: {len(df)} 条数据")
                except Exception as e:
                    print(f"[Step3] ✗ 读取 {platform} 文件失败: {e}")
                    continue
        
        print(f"\n[Step3] 汇总Excel已生成: {merged_file}")
        print(f"[Step3] 共合并 {len(platform_files)} 个平台的数据")
        return merged_file
        
    except Exception as e:
        print(f"[Step3] 合并Excel时出错: {e}")
        import traceback
        traceback.print_exc()
        return None


def run_automated_workflow(
    zol_start_page: int = 1,
    zol_end_page: int = 3,
    excel_file_path: Optional[str] = None,
    xhs_limit: int = 200,
    xhs_headless: bool = True,
    run_xhs: bool = True,
    skip_step1: bool = False,
    run_weibo: bool = True,
    weibo_limit: int = 200,
    weibo_headless: bool = False,
    run_douban: bool = True,
    douban_headless: bool = False,
    run_zhihu: bool = True,
    zhihu_headless: bool = False,
    run_bilibili: bool = True,
):
    """
    自动化工作流：step1 中关村爬虫 -> step2 豆瓣/小红书/微博/知乎/B站爬虫
    
    参数:
        zol_start_page: 中关村爬虫起始页码
        zol_end_page: 中关村爬虫结束页码
        excel_file_path: Excel文件路径，如果为None则使用step1生成的最新文件
        xhs_limit: 小红书爬虫每个关键词的采集数量上限
        xhs_headless: 小红书爬虫是否使用无头模式
        run_xhs: 是否运行小红书爬虫
        skip_step1: 是否跳过step1，直接使用现有的Excel文件
        run_weibo: 是否在Step2阶段同时运行微博爬虫
        weibo_limit: 微博爬虫每个关键词采集数量上限
        weibo_headless: 微博爬虫是否使用无头模式
        run_douban: 是否在Step2阶段同时运行豆瓣爬虫
        douban_headless: 豆瓣爬虫是否使用无头模式
        run_zhihu: 是否在Step2阶段同时运行知乎爬虫
        zhihu_headless: 知乎爬虫是否使用无头模式
        run_bilibili: 是否在Step2阶段同时运行B站爬虫
    """
    excel_path = excel_file_path
    
    # Step 1: 运行中关村爬虫（如果未跳过）
    if not skip_step1:
        print("=" * 60)
        print("Step 1: 开始运行中关村爬虫...")
        print("=" * 60)
        
        products = crawl_zol_products(start_page=zol_start_page, end_page=zol_end_page)
        df = pd.DataFrame(products)
        print("\n数据预览：")
        print(df)
        
        # 导出到Excel
        excel_path = export_to_excel(df, filename=None, formatted=True)
    else:
        print("=" * 60)
        print("Step 1: 跳过中关村爬虫，使用现有Excel文件")
        print("=" * 60)
        
        if excel_path is None:
            # 查找最新的中关村Excel文件
            import glob
            excel_files = glob.glob("中关村在线产品信息_*.xlsx")
            if excel_files:
                excel_files.sort(reverse=True)
                excel_path = excel_files[0]
                print(f"自动找到最新的Excel文件: {excel_path}")
            else:
                raise FileNotFoundError("未找到中关村爬虫输出的Excel文件，请先运行step1或指定文件路径")
    
    # 读取手机型号列表
    print("\n" + "=" * 60)
    print("读取Excel文件中的手机型号...")
    print("=" * 60)
    phone_models = read_phone_models_from_excel(excel_path)
    
    if not phone_models:
        print("警告: 未找到任何手机型号，无法继续执行step2")
        return
    
    print(f"\n找到 {len(phone_models)} 个手机型号，将依次进行豆瓣/小红书/微博/知乎爬取:")
    for i, model in enumerate(phone_models, 1):
        print(f"  {i}. {model}")
    
    # Step 2: 豆瓣 / 小红书 / 微博 / 知乎 / B站 并行启动
    print("\n" + "=" * 60)
    print("Step 2: 开始运行豆瓣 / 小红书 / 微博 / 知乎 / B站爬虫（并行启动，可独立选择有头/无头）...")
    print("=" * 60)

    xhs_all_data: List[tuple] = []
    douban_success_cnt = 0
    zhihu_success_cnt = 0

    def run_xhs_workflow():
        for idx, phone_model in enumerate(phone_models, 1):
            print(f"\n{'=' * 60}")
            print(f"[XHS] 正在处理第 {idx}/{len(phone_models)} 个手机型号: {phone_model}")
            print(f"{'=' * 60}")
            
            try:
                scraper = XiaoHongShuScraper(
                    keyword=phone_model,
                    limit=xhs_limit,
                    headless=xhs_headless,
                    timeout_sec=25,
                    delay_min=1.2,
                    delay_max=2.4,
                    output_path=None,  # 使用默认路径，会自动生成文件名
                    proxy=None,
                )
                
                records = scraper.scrape()
                scraper.export_excel(records)
                xhs_all_data.append((phone_model, records))
                print(f"[XHS] ✓ 手机型号 '{phone_model}' 的小红书数据爬取完成，共 {len(records)} 条记录")
                
                if idx < len(phone_models):
                    delay = random.uniform(3, 5)
                    print(f"[XHS] 等待 {delay:.1f} 秒后处理下一个关键词...")
                    time.sleep(delay)
                    
            except Exception as e:
                print(f"[XHS] ✗ 处理手机型号 '{phone_model}' 时出错: {e}")
                print("[XHS] 继续处理下一个手机型号...")
                continue

    def run_weibo_workflow():
        if not run_weibo:
            return
        print("[Weibo] 已启用微博爬虫，开始批量执行...")
        runner = WeiboRunner(limit=weibo_limit, headless=weibo_headless)
        try:
            success_cnt = runner.run_keywords(phone_models)
            print(f"[Weibo] 成功爬取 {success_cnt}/{len(phone_models)} 个关键词")
        finally:
            runner.close()

    def run_douban_thread():
        nonlocal douban_success_cnt
        if not run_douban:
            return
        print("[Douban] 已启用豆瓣爬虫，开始批量执行...")
        douban_success_cnt = run_douban_workflow(phone_models, headless=douban_headless)
        print(f"[Douban] 成功爬取 {douban_success_cnt}/{len(phone_models)} 个关键词")

    def run_zhihu_workflow():
        nonlocal zhihu_success_cnt
        if not run_zhihu:
            return
        print("[Zhihu] 已启用知乎爬虫，开始批量执行...")
        zhihu_success_cnt = zhihu_run_keywords(phone_models, headless=zhihu_headless)
        print(f"[Zhihu] 成功爬取 {zhihu_success_cnt}/{len(phone_models)} 个关键词")

    def run_bilibili_workflow():
        if not run_bilibili:
            return
        print("[Bilibili] 已启用B站爬虫，开始批量执行...")
        try:
            bili_run_keywords(phone_models)
        except Exception as e:
            print(f"[Bilibili] 执行出错: {e}")

    threads = []
    if run_douban:
        douban_thread = threading.Thread(target=run_douban_thread, name="douban-thread", daemon=False)
        threads.append(douban_thread)
        douban_thread.start()
    else:
        print("[Douban] 已禁用豆瓣爬虫，跳过。")

    if run_zhihu:
        zhihu_thread = threading.Thread(target=run_zhihu_workflow, name="zhihu-thread", daemon=False)
        threads.append(zhihu_thread)
        zhihu_thread.start()
    else:
        print("[Zhihu] 已禁用知乎爬虫，跳过。")

    if run_xhs:
        xhs_thread = threading.Thread(target=run_xhs_workflow, name="xhs-thread", daemon=False)
        threads.append(xhs_thread)
        xhs_thread.start()
    else:
        print("[XHS] 已禁用小红书爬虫，跳过。")

    if run_weibo:
        weibo_thread = threading.Thread(target=run_weibo_workflow, name="weibo-thread", daemon=False)
        threads.append(weibo_thread)
        weibo_thread.start()

    if run_bilibili:
        bilibili_thread = threading.Thread(target=run_bilibili_workflow, name="bilibili-thread", daemon=False)
        threads.append(bilibili_thread)
        bilibili_thread.start()

    for t in threads:
        t.join()
    
    # 关键词整合导出
    if run_xhs and xhs_all_data:
        xhs_out_dir = os.path.abspath(XHS_OUTPUT_DIR)
        xhs_save_combined_excel(xhs_all_data, xhs_out_dir)
    if run_weibo:
        # WeiboRunner 内部已导出整合文件到 WEIBO_OUTPUT_DIR
        pass
    if run_bilibili:
        # bilibili.main 内部已实时写入各类 CSV/JSON 文件
        pass
    
    # Step 3: 合并所有平台的整合Excel文件
    print("\n" + "=" * 60)
    print("Step 3: 开始合并所有平台的整合Excel文件...")
    print("=" * 60)
    step3_merge_all_excel(
        run_xhs=run_xhs,
        run_weibo=run_weibo,
        run_douban=run_douban,
        run_zhihu=run_zhihu,
        run_bilibili=run_bilibili,
    )
    
    print("\n" + "=" * 60)
    print("自动化工作流执行完成！")
    print("=" * 60)


import argparse
import os
import random
import re
import sys
import time
import traceback
import threading
import pickle
import csv
from dataclasses import dataclass, asdict
from typing import Dict
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from xml.etree import ElementTree as ET
from xml.etree import ElementTree as ET

from dotenv import load_dotenv
from playwright.sync_api import Browser, BrowserContext, Page, sync_playwright

try:
    from rich.console import Console
    from rich.progress import track
    console = Console()
except Exception:  # rich is optional
    class _Dummy:
        def print(self, *args, **kwargs):
            print(*args)
    console = _Dummy()
    def track(iterable, description=""):
        return iterable



SEARCH_URL_TEMPLATE = "https://www.xiaohongshu.com/search_result?keyword={keyword}"
NOTE_URL_PREFIX = "https://www.xiaohongshu.com/explore/"
STORAGE_STATE_FILE = "storage_state.json"
DEFAULT_OUTPUT_DIR = "out"
XHS_OUTPUT_DIR = os.path.join(DEFAULT_OUTPUT_DIR, "xhs")
WEIBO_OUTPUT_DIR = os.path.join(DEFAULT_OUTPUT_DIR, "weibo")
DOUBAN_OUTPUT_DIR = os.path.join(DEFAULT_OUTPUT_DIR, "douban")
ZHIHU_OUTPUT_DIR = os.path.join(DEFAULT_OUTPUT_DIR, "zhihu")
BILIBILI_OUTPUT_DIR = os.path.join(DEFAULT_OUTPUT_DIR, "bilibili")
BILIBILI_OUTPUT_DIR = os.path.join(DEFAULT_OUTPUT_DIR, "bilibili")

# ======================
# 知乎爬虫内联实现（由 zhihu.py 嵌入）
# ======================
ZHIHU_PAGE_LIMIT = 10
ZHIHU_COOKIE_FILE = os.path.join(ZHIHU_OUTPUT_DIR, "zhihu_dom_cookies.pkl")
ZHIHU_DOM_CSV = os.path.join(ZHIHU_OUTPUT_DIR, "zhihu_dom_all.csv")
ZHIHU_CHECKPOINT_FILE = os.path.join(ZHIHU_OUTPUT_DIR, "checkpoint.txt")
ZHIHU_HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
    "referer": "https://www.zhihu.com/",
}
ZHIHU_ALL_CSV_FD = None
ZHIHU_ALL_CSV_WRITER = None

# ======================
# B站爬虫内联实现（由 bilibili.py 嵌入）
# ======================
BILI_PAGE_PER_KEY = 1
BILI_COMMENTS_PAGES = 10
BILI_COOKIE_DIR = BILIBILI_OUTPUT_DIR
BILI_COOKIE_FILE = os.path.join(BILI_COOKIE_DIR, "bili_cookies.json")
BILI_CHECKPOINT = os.path.join(BILIBILI_OUTPUT_DIR, "checkpoint.json")
BILI_SESSION = requests.Session()
BILI_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
)
BILI_SESSION.headers.update(
    {"User-Agent": BILI_UA, "Referer": "https://search.bilibili.com/", "Origin": "https://search.bilibili.com"}
)

def bili_next_filename(base: str, ext: str = "csv"):
    """生成带自增序号的文件名，保证目录存在并对名称做安全清洗。"""
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    safe_base = _sanitize_filename(base)
    cnt = 1
    while os.path.exists(os.path.join(BILIBILI_OUTPUT_DIR, f"{safe_base}_{cnt:03d}.{ext}")):
        cnt += 1
    return os.path.join(BILIBILI_OUTPUT_DIR, f"{safe_base}_{cnt:03d}.{ext}")


def bili_append_csv(base: str, row: dict):
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    path = os.path.join(BILIBILI_OUTPUT_DIR, f"{base}.csv")
    header = not os.path.exists(path)
    try:
        pd.DataFrame([row]).to_csv(path, mode="a", header=header, index=False, encoding="utf-8-sig")
    except Exception as e:
        print(f"[Bilibili] 写入失败 {path}: {e}")
        traceback.print_exc()


def bili_save_json_once(base: str, data):
    """保存一次搜索结果 JSON（包含目录创建与文件名清洗）。"""
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    safe_base = _sanitize_filename(base)
    fn = bili_next_filename(safe_base, "json")
    with open(fn, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return fn


def bili_load_cookies_to_session():
    if not os.path.exists(BILI_COOKIE_FILE):
        return False
    with open(BILI_COOKIE_FILE, "r", encoding="utf-8") as f:
        cookies = json.load(f)
    BILI_SESSION.cookies.update({c["name"]: c["value"] for c in cookies})
    return True


def bili_save_cookies_from_driver(driver):
    with open(BILI_COOKIE_FILE, "w", encoding="utf-8") as f:
        json.dump(driver.get_cookies(), f, ensure_ascii=False, indent=2)


def bili_open_browser_for_login():
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.bilibili.com")
    input(">>> 请扫码登录，完成后按 Enter 继续...")
    bili_save_cookies_from_driver(driver)
    driver.quit()


def bili_anti_spider_intervention(url: str):
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.bilibili.com")
    time.sleep(1)
    driver.get(url)
    input(">>> 处理完成后按 Enter 继续...")
    bili_save_cookies_from_driver(driver)
    driver.quit()


def bili_search_videos(kw: str, pn: int):
    url = "https://api.bilibili.com/x/web-interface/search/type"
    params = {"search_type": "video", "keyword": kw, "page": pn, "pagesize": 20}
    r = BILI_SESSION.get(url, params=params)
    if r.status_code != 200:
        raise RuntimeError(f"http {r.status_code}")
    ret = r.json()
    if ret["code"] != 0:
        raise RuntimeError(f"api code={ret['code']} msg={ret.get('message')}")
    return ret


def bili_get_video_detail(bv: str):
    url = f"https://api.bilibili.com/x/web-interface/view?bvid={bv}"
    return BILI_SESSION.get(url).json()


def bili_get_comments(oid: str, page: int):
    url = "https://api.bilibili.com/x/v2/reply"
    params = {"type": 1, "oid": oid, "pn": page, "sort": 0}
    return BILI_SESSION.get(url, params=params).json()


def bili_get_danmu(cid: str):
    url = f"https://comment.bilibili.com/{cid}.xml"
    r = BILI_SESSION.get(url)
    r.encoding = "utf-8"
    root = ET.fromstring(r.text)
    return [{"time": float(d.get("p").split(",")[0]), "text": d.text} for d in root.iter("d")]


def bili_load_checkpoint():
    if not os.path.exists(BILI_CHECKPOINT):
        return set()
    try:
        return set(json.load(open(BILI_CHECKPOINT, "r", encoding="utf-8")).get("done", []))
    except Exception:
        return set()


def bili_save_checkpoint(bv: str):
    done = bili_load_checkpoint()
    done.add(bv)
    with open(BILI_CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump({"done": list(done)}, f)


def bili_process_one_video(v: dict, keyword: str, acc: dict):
    """处理单个视频并把行数据追加到缓存，便于后续导出 Excel。"""
    bv = v.get("bvid")
    if not bv:
        return

    def clean_title(title):
        return re.sub(r"<[^>]+>", "", title) if title else ""

    search_title = clean_title(v.get("title", ""))
    v_row = {
        "search_keyword": keyword,
        "bv": bv,
        "title": search_title,
        "author": v.get("author", ""),
        "arcurl": v.get("arcurl", ""),
        "duration": v.get("duration", ""),
        "play": v.get("play", 0),
        "pubdate": v.get("senddate", 0),
        "description": v.get("description", ""),
        "video_review": v.get("video_review", 0),
        "favorites": v.get("favorites", 0),
        "tag": v.get("tag", ""),
        "review": v.get("review", 0),
        "mid": v.get("mid", ""),
        "typename": v.get("typename", ""),
        "pic": v.get("pic", ""),
        "is_pay": v.get("is_pay", 0),
        "is_union_video": v.get("is_union_video", 0),
        "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    bili_append_csv("search_list", v_row)
    acc["search"].append(v_row)

    detail = bili_get_video_detail(bv)
    if detail.get("code") != 0:
        print(f"[Bilibili] 详情接口返回非0，跳过：{detail.get('message')}")
        return
    d = detail.get("data", {})
    detail_title = d.get("title", "")
    detail_row = {
        "search_keyword": keyword,
        "bv": bv,
        "aid": d.get("aid", ""),
        "cid": d.get("cid", ""),
        "view": d.get("stat", {}).get("view", 0),
        "danmaku": d.get("stat", {}).get("danmaku", 0),
        "reply": d.get("stat", {}).get("reply", 0),
        "favorite": d.get("stat", {}).get("favorite", 0),
        "coin": d.get("stat", {}).get("coin", 0),
        "share": d.get("stat", {}).get("share", 0),
        "like": d.get("stat", {}).get("like", 0),
        "dislike": d.get("stat", {}).get("dislike", 0),
        "now_rank": d.get("stat", {}).get("now_rank", 0),
        "his_rank": d.get("stat", {}).get("his_rank", 0),
        "evaluation": d.get("stat", {}).get("evaluation", ""),
        "argue_msg": d.get("stat", {}).get("argue_msg", ""),
        "tname": d.get("tname", ""),
        "pubdate": d.get("pubdate", 0),
        "ctime": d.get("ctime", 0),
        "title": detail_title,
        "desc": d.get("desc", ""),
        "dynamic": d.get("dynamic", ""),
        "videos": d.get("videos", 1),
        "tid": d.get("tid", ""),
        "copyright": d.get("copyright", 0),
        "owner_mid": d.get("owner", {}).get("mid", ""),
        "owner_name": d.get("owner", {}).get("name", ""),
        "owner_face": d.get("owner", {}).get("face", ""),
        "pages": len(d.get("pages", [])),
        "subtitle_count": len(d.get("subtitle", {}).get("list", [])),
        "is_chargeable_season": d.get("is_chargeable_season", 0),
        "is_blooper": d.get("is_blooper", 0),
        "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    bili_append_csv("video_detail", detail_row)
    acc["detail"].append(detail_row)

    oid = d.get("aid")
    cid = d.get("cid")
    if not oid:
        print("[Bilibili] 无法获取aid，跳过评论/弹幕")
        return

    page = 1
    comment_count = 0
    while page <= BILI_COMMENTS_PAGES:
        ret = bili_get_comments(oid, page)
        if ret.get("code") != 0 or not ret.get("data", {}).get("replies"):
            break
        for r in ret["data"]["replies"]:
            comment_row = {
                "search_keyword": keyword,
                "video_title": detail_title,
                "bv": bv,
                "mid": r.get("mid", ""),
                "uname": r.get("member", {}).get("uname", ""),
                "sex": r.get("member", {}).get("sex", ""),
                "sign": r.get("member", {}).get("sign", ""),
                "level": r.get("member", {}).get("level_info", {}).get("current_level", 0),
                "like": r.get("like", 0),
                "rcount": r.get("rcount", 0),
                "count": r.get("count", 0),
                "ctime": r.get("ctime", 0),
                "content": r.get("content", {}).get("message", ""),
                "root": r.get("root", 0),
                "parent": r.get("parent", 0),
                "dialog": r.get("dialog", 0),
                "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            bili_append_csv("comment", comment_row)
            acc["comment"].append(comment_row)
            comment_count += 1
        page += 1
        time.sleep(random.uniform(1, 2))
    print(f"[Bilibili] 获取到 {comment_count} 条评论")

    if cid:
        try:
            danmus = bili_get_danmu(cid)
            for dm in danmus:
                danmu_row = {
                    "search_keyword": keyword,
                    "video_title": detail_title,
                    "bv": bv,
                    "time": dm["time"],
                    "text": dm["text"],
                    "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                bili_append_csv("danmu", danmu_row)
                acc["danmu"].append(danmu_row)
            print(f"[Bilibili] 获取到 {len(danmus)} 条弹幕")
        except Exception as e:
            print(f"[Bilibili] 获取弹幕失败: {e}")


def bili_export_excel(keyword: str, data: dict, filename: Optional[str] = None) -> Optional[str]:
    """将单关键词或汇总数据导出为 Excel（含多表）。"""
    if not any(data.values()):
        print(f"[Bilibili] 关键词 {keyword} 无数据，跳过导出")
        return None
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    safe_kw = _sanitize_filename(keyword) or "keyword"
    filename = filename or os.path.join(BILIBILI_OUTPUT_DIR, f"bilibili_{safe_kw}.xlsx")
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        if data.get("search"):
            pd.DataFrame(data["search"]).to_excel(writer, sheet_name="search_list", index=False)
        if data.get("detail"):
            pd.DataFrame(data["detail"]).to_excel(writer, sheet_name="video_detail", index=False)
        if data.get("comment"):
            pd.DataFrame(data["comment"]).to_excel(writer, sheet_name="comment", index=False)
        if data.get("danmu"):
            pd.DataFrame(data["danmu"]).to_excel(writer, sheet_name="danmu", index=False)
    print(f"[Bilibili] 已导出 Excel -> {filename}")
    return filename


def bili_run_keywords(keywords: List[str]):
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    os.makedirs(BILI_COOKIE_DIR, exist_ok=True)
    print("[Bilibili] 目录准备完成")
    print("[Bilibili] 开始加载 cookie...")
    if not bili_load_cookies_to_session():
        print("[Bilibili] 无 cookie，打开浏览器登录...")
        bili_open_browser_for_login()
        bili_load_cookies_to_session()
    print("[Bilibili] cookie 已加载，开始搜索...")
    done = bili_load_checkpoint()

    combined = {"search": [], "detail": [], "comment": [], "danmu": []}

    for kw in keywords:
        per_kw = {"search": [], "detail": [], "comment": [], "danmu": []}
        for pn in range(1, BILI_PAGE_PER_KEY + 1):
            print(f"[Bilibili] 搜索 【{kw}】第 {pn} 页")
            try:
                ret = bili_search_videos(kw, pn)
            except Exception as e:
                print(f"[Bilibili] 搜索失败 {kw}: {e}")
                continue
            videos = ret.get("data", {}).get("result", [])
            if not videos:
                print("[Bilibili] 没有找到视频数据")
                continue

            bili_save_json_once(f"search_{kw}", ret)
            for v in videos:
                if not isinstance(v, dict):
                    continue
                bv = v.get("bvid")
                if not bv:
                    continue
                if bv in done:
                    print(f"[Bilibili] 跳过已处理的视频: {bv}")
                    continue
                try:
                    print(f"[Bilibili] 开始处理: {bv}")
                    bili_process_one_video(v, kw, per_kw)
                    bili_save_checkpoint(bv)
                    print(f"[Bilibili] 完成处理: {bv}")
                except Exception as e:
                    print(f"[Bilibili] {bv} 处理失败：{e}")
                    print("[Bilibili] 准备人工干预...")
                    bili_anti_spider_intervention(f"https://www.bilibili.com/video/{bv}")
                    bili_load_cookies_to_session()
                time.sleep(random.uniform(5, 10))

        # 每个关键词导出独立 Excel
        bili_export_excel(kw, per_kw)
        # 汇总累加
        for k in combined.keys():
            combined[k].extend(per_kw[k])

    # 全部关键词完成后导出整合 Excel
    bili_export_excel("all", combined, os.path.join(BILIBILI_OUTPUT_DIR, "bilibili_all.xlsx"))

# ======================
# B站爬虫内联实现（由 bilibili.py 嵌入）
# ======================
BILI_PAGE_PER_KEY = 1
BILI_COMMENTS_PAGES = 10
BILI_COOKIE_DIR = BILIBILI_OUTPUT_DIR
BILI_COOKIE_FILE = os.path.join(BILI_COOKIE_DIR, "bili_cookies.json")
BILI_CHECKPOINT = os.path.join(BILIBILI_OUTPUT_DIR, "checkpoint.json")
BILI_SESSION = requests.Session()
BILI_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
)
BILI_SESSION.headers.update(
    {"User-Agent": BILI_UA, "Referer": "https://search.bilibili.com/", "Origin": "https://search.bilibili.com"}
)

def bili_next_filename(base: str, ext: str = "csv"):
    """生成带自增序号的文件名，保证目录存在并对名称做安全清洗。"""
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    safe_base = _sanitize_filename(base)
    cnt = 1
    while os.path.exists(os.path.join(BILIBILI_OUTPUT_DIR, f"{safe_base}_{cnt:03d}.{ext}")):
        cnt += 1
    return os.path.join(BILIBILI_OUTPUT_DIR, f"{safe_base}_{cnt:03d}.{ext}")


def bili_append_csv(base: str, row: dict):
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    path = os.path.join(BILIBILI_OUTPUT_DIR, f"{base}.csv")
    header = not os.path.exists(path)
    try:
        pd.DataFrame([row]).to_csv(path, mode="a", header=header, index=False, encoding="utf-8-sig")
    except Exception as e:
        print(f"[Bilibili] 写入失败 {path}: {e}")
        traceback.print_exc()


def bili_save_json_once(base: str, data):
    """保存一次搜索结果 JSON（包含目录创建与文件名清洗）。"""
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    safe_base = _sanitize_filename(base)
    fn = bili_next_filename(safe_base, "json")
    with open(fn, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return fn


def bili_load_cookies_to_session():
    if not os.path.exists(BILI_COOKIE_FILE):
        return False
    with open(BILI_COOKIE_FILE, "r", encoding="utf-8") as f:
        cookies = json.load(f)
    BILI_SESSION.cookies.update({c["name"]: c["value"] for c in cookies})
    return True


def bili_save_cookies_from_driver(driver):
    with open(BILI_COOKIE_FILE, "w", encoding="utf-8") as f:
        json.dump(driver.get_cookies(), f, ensure_ascii=False, indent=2)


def bili_open_browser_for_login():
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.bilibili.com")
    input(">>> 请扫码登录，完成后按 Enter 继续...")
    bili_save_cookies_from_driver(driver)
    driver.quit()


def bili_anti_spider_intervention(url: str):
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.bilibili.com")
    time.sleep(1)
    driver.get(url)
    input(">>> 处理完成后按 Enter 继续...")
    bili_save_cookies_from_driver(driver)
    driver.quit()


def bili_search_videos(kw: str, pn: int):
    url = "https://api.bilibili.com/x/web-interface/search/type"
    params = {"search_type": "video", "keyword": kw, "page": pn, "pagesize": 20}
    r = BILI_SESSION.get(url, params=params)
    if r.status_code != 200:
        raise RuntimeError(f"http {r.status_code}")
    ret = r.json()
    if ret["code"] != 0:
        raise RuntimeError(f"api code={ret['code']} msg={ret.get('message')}")
    return ret


def bili_get_video_detail(bv: str):
    url = f"https://api.bilibili.com/x/web-interface/view?bvid={bv}"
    return BILI_SESSION.get(url).json()


def bili_get_comments(oid: str, page: int):
    url = "https://api.bilibili.com/x/v2/reply"
    params = {"type": 1, "oid": oid, "pn": page, "sort": 0}
    return BILI_SESSION.get(url, params=params).json()


def bili_get_danmu(cid: str):
    url = f"https://comment.bilibili.com/{cid}.xml"
    r = BILI_SESSION.get(url)
    r.encoding = "utf-8"
    root = ET.fromstring(r.text)
    return [{"time": float(d.get("p").split(",")[0]), "text": d.text} for d in root.iter("d")]


def bili_load_checkpoint():
    if not os.path.exists(BILI_CHECKPOINT):
        return set()
    try:
        return set(json.load(open(BILI_CHECKPOINT, "r", encoding="utf-8")).get("done", []))
    except Exception:
        return set()


def bili_save_checkpoint(bv: str):
    done = bili_load_checkpoint()
    done.add(bv)
    with open(BILI_CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump({"done": list(done)}, f)


def bili_process_one_video(v: dict, keyword: str):
    bv = v.get("bvid")
    if not bv:
        return

    def clean_title(title):
        return re.sub(r"<[^>]+>", "", title) if title else ""

    search_title = clean_title(v.get("title", ""))
    v_row = {
        "search_keyword": keyword,
        "bv": bv,
        "title": search_title,
        "author": v.get("author", ""),
        "arcurl": v.get("arcurl", ""),
        "duration": v.get("duration", ""),
        "play": v.get("play", 0),
        "pubdate": v.get("senddate", 0),
        "description": v.get("description", ""),
        "video_review": v.get("video_review", 0),
        "favorites": v.get("favorites", 0),
        "tag": v.get("tag", ""),
        "review": v.get("review", 0),
        "mid": v.get("mid", ""),
        "typename": v.get("typename", ""),
        "pic": v.get("pic", ""),
        "is_pay": v.get("is_pay", 0),
        "is_union_video": v.get("is_union_video", 0),
        "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    bili_append_csv("search_list", v_row)

    detail = bili_get_video_detail(bv)
    if detail.get("code") != 0:
        print(f"[Bilibili] 详情接口返回非0，跳过：{detail.get('message')}")
        return
    d = detail.get("data", {})
    detail_title = d.get("title", "")
    detail_row = {
        "search_keyword": keyword,
        "bv": bv,
        "aid": d.get("aid", ""),
        "cid": d.get("cid", ""),
        "view": d.get("stat", {}).get("view", 0),
        "danmaku": d.get("stat", {}).get("danmaku", 0),
        "reply": d.get("stat", {}).get("reply", 0),
        "favorite": d.get("stat", {}).get("favorite", 0),
        "coin": d.get("stat", {}).get("coin", 0),
        "share": d.get("stat", {}).get("share", 0),
        "like": d.get("stat", {}).get("like", 0),
        "dislike": d.get("stat", {}).get("dislike", 0),
        "now_rank": d.get("stat", {}).get("now_rank", 0),
        "his_rank": d.get("stat", {}).get("his_rank", 0),
        "evaluation": d.get("stat", {}).get("evaluation", ""),
        "argue_msg": d.get("stat", {}).get("argue_msg", ""),
        "tname": d.get("tname", ""),
        "pubdate": d.get("pubdate", 0),
        "ctime": d.get("ctime", 0),
        "title": detail_title,
        "desc": d.get("desc", ""),
        "dynamic": d.get("dynamic", ""),
        "videos": d.get("videos", 1),
        "tid": d.get("tid", ""),
        "copyright": d.get("copyright", 0),
        "owner_mid": d.get("owner", {}).get("mid", ""),
        "owner_name": d.get("owner", {}).get("name", ""),
        "owner_face": d.get("owner", {}).get("face", ""),
        "pages": len(d.get("pages", [])),
        "subtitle_count": len(d.get("subtitle", {}).get("list", [])),
        "is_chargeable_season": d.get("is_chargeable_season", 0),
        "is_blooper": d.get("is_blooper", 0),
        "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    bili_append_csv("video_detail", detail_row)

    oid = d.get("aid")
    cid = d.get("cid")
    if not oid:
        print("[Bilibili] 无法获取aid，跳过评论/弹幕")
        return

    page = 1
    comment_count = 0
    while page <= BILI_COMMENTS_PAGES:
        ret = bili_get_comments(oid, page)
        if ret.get("code") != 0 or not ret.get("data", {}).get("replies"):
            break
        for r in ret["data"]["replies"]:
            comment_row = {
                "search_keyword": keyword,
                "video_title": detail_title,
                "bv": bv,
                "mid": r.get("mid", ""),
                "uname": r.get("member", {}).get("uname", ""),
                "sex": r.get("member", {}).get("sex", ""),
                "sign": r.get("member", {}).get("sign", ""),
                "level": r.get("member", {}).get("level_info", {}).get("current_level", 0),
                "like": r.get("like", 0),
                "rcount": r.get("rcount", 0),
                "count": r.get("count", 0),
                "ctime": r.get("ctime", 0),
                "content": r.get("content", {}).get("message", ""),
                "root": r.get("root", 0),
                "parent": r.get("parent", 0),
                "dialog": r.get("dialog", 0),
                "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            bili_append_csv("comment", comment_row)
            comment_count += 1
        page += 1
        time.sleep(random.uniform(1, 2))
    print(f"[Bilibili] 获取到 {comment_count} 条评论")

    if cid:
        try:
            danmus = bili_get_danmu(cid)
            for dm in danmus:
                bili_append_csv(
                    "danmu",
                    {
                        "search_keyword": keyword,
                        "video_title": detail_title,
                        "bv": bv,
                        "time": dm["time"],
                        "text": dm["text"],
                        "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    },
                )
            print(f"[Bilibili] 获取到 {len(danmus)} 条弹幕")
        except Exception as e:
            print(f"[Bilibili] 获取弹幕失败: {e}")


def bili_run_keywords(keywords: List[str]):
    os.makedirs(BILIBILI_OUTPUT_DIR, exist_ok=True)
    os.makedirs(BILI_COOKIE_DIR, exist_ok=True)
    print("[Bilibili] 目录准备完成")
    print("[Bilibili] 开始加载 cookie...")
    if not bili_load_cookies_to_session():
        print("[Bilibili] 无 cookie，打开浏览器登录...")
        bili_open_browser_for_login()
        bili_load_cookies_to_session()
    print("[Bilibili] cookie 已加载，开始搜索...")
    done = bili_load_checkpoint()

    for kw in keywords:
        for pn in range(1, BILI_PAGE_PER_KEY + 1):
            print(f"[Bilibili] 搜索 【{kw}】第 {pn} 页")
            try:
                ret = bili_search_videos(kw, pn)
            except Exception as e:
                print(f"[Bilibili] 搜索失败 {kw}: {e}")
                continue
            videos = ret.get("data", {}).get("result", [])
            if not videos:
                print("[Bilibili] 没有找到视频数据")
                continue

            bili_save_json_once(f"search_{kw}", ret)
            for v in videos:
                if not isinstance(v, dict):
                    continue
                bv = v.get("bvid")
                if not bv:
                    continue
                if bv in done:
                    print(f"[Bilibili] 跳过已处理的视频: {bv}")
                    continue
                try:
                    print(f"[Bilibili] 开始处理: {bv}")
                    bili_process_one_video(v, kw)
                    bili_save_checkpoint(bv)
                    print(f"[Bilibili] 完成处理: {bv}")
                except Exception as e:
                    print(f"[Bilibili] {bv} 处理失败：{e}")
                    print("[Bilibili] 准备人工干预...")
                    bili_anti_spider_intervention(f"https://www.bilibili.com/video/{bv}")
                    bili_load_cookies_to_session()
                time.sleep(random.uniform(5, 10))


def zhihu_ensure_dir():
    os.makedirs(ZHIHU_OUTPUT_DIR, exist_ok=True)


def zhihu_save_driver_cookies(driver):
    with open(ZHIHU_COOKIE_FILE, "wb") as f:
        pickle.dump(driver.get_cookies(), f)
    print("[Zhihu] Cookies 已保存 ->", ZHIHU_COOKIE_FILE)


def zhihu_load_driver_cookies(driver) -> bool:
    if not os.path.exists(ZHIHU_COOKIE_FILE):
        return False
    driver.get("https://www.zhihu.com")
    driver.delete_all_cookies()
    for c in pickle.load(open(ZHIHU_COOKIE_FILE, "rb")):
        c.pop("sameSite", None)
        driver.add_cookie(c)
    print("[Zhihu] Cookies 已加载 <-", ZHIHU_COOKIE_FILE)
    return True


def zhihu_open_csv_once():
    global ZHIHU_ALL_CSV_FD, ZHIHU_ALL_CSV_WRITER
    if ZHIHU_ALL_CSV_FD:
        return ZHIHU_ALL_CSV_WRITER
    header = not os.path.exists(ZHIHU_DOM_CSV)
    ZHIHU_ALL_CSV_FD = open(ZHIHU_DOM_CSV, "a", encoding="utf-8-sig", newline="")
    ZHIHU_ALL_CSV_WRITER = csv.DictWriter(
        ZHIHU_ALL_CSV_FD,
        fieldnames=[
            "关键词",
            "回答ID",
            "问题ID",
            "问题标题",
            "回答链接",
            "回答正文",
            "回答创建时间",
            "回答更新时间",
            "点赞数",
            "评论数",
            "作者ID",
            "作者昵称",
            "作者性别",
            "作者头像",
            "作者主页",
            "作者粉丝数",
            "评论",
            "原始JSON",
            "抓取时间",
        ],
    )
    if header:
        ZHIHU_ALL_CSV_WRITER.writeheader()
    return ZHIHU_ALL_CSV_WRITER


def zhihu_append_one_row(row: dict):
    writer = zhihu_open_csv_once()
    writer.writerow(row)
    ZHIHU_ALL_CSV_FD.flush()


def zhihu_close_csv():
    global ZHIHU_ALL_CSV_FD, ZHIHU_ALL_CSV_WRITER
    if ZHIHU_ALL_CSV_FD:
        try:
            ZHIHU_ALL_CSV_FD.close()
        except Exception:
            pass
    ZHIHU_ALL_CSV_FD = None
    ZHIHU_ALL_CSV_WRITER = None


def zhihu_load_checkpoint() -> set:
    if not os.path.exists(ZHIHU_CHECKPOINT_FILE):
        return set()
    with open(ZHIHU_CHECKPOINT_FILE, encoding="utf-8") as f:
        return {line.strip() for line in f if line.strip()}


def zhihu_save_checkpoint(keyword: str):
    with open(ZHIHU_CHECKPOINT_FILE, "a", encoding="utf-8") as f:
        f.write(keyword + "\n")


def zhihu_login_by_selenium(driver):
    print("[Zhihu] 打开登录页，请扫码…")
    driver.get("https://www.zhihu.com/signin")
    try:
        WebDriverWait(driver, 25).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".Qrcode-signInStep"))
        )
    except Exception:
        print("[Zhihu] 没等到二维码，继续往下走……")
    WebDriverWait(driver, 60).until(
        lambda d: d.current_url
        in {"https://www.zhihu.com/", "https://www.zhihu.com", "https://www.zhihu.com/explore"}
    )
    time.sleep(3)
    zhihu_save_driver_cookies(driver)
    print("[Zhihu] 登录成功，Cookies 已保存")


def zhihu_search_by_keyword_dom(keyword: str, driver: webdriver.Chrome):
    url = f"https://www.zhihu.com/search?type=content&q={keyword}"
    driver.get(url)
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.SearchResult-Card, div[data-za-detail-view-path-module='AnswerItem']"))
        )
    except Exception:
        pass
    time.sleep(2)

    seen_ids, page, new_in_last = set(), 0, 1
    all_cards = []

    while new_in_last and page < ZHIHU_PAGE_LIMIT:
        page += 1
        new_in_last = 0

        for _ in range(6):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)

        soup = BeautifulSoup(driver.page_source, "lxml")
        # 兼容新版/旧版结构
        cards = soup.select('div[data-za-detail-view-path-module="AnswerItem"]')
        if not cards:
            cards = soup.select("div.SearchResult-Card")
        if not cards:
            cards = soup.select("div.Card, div.TopstoryItem")

        for c in cards:
            aid_tag = c.select_one('a[href*="/answer/"]')
            if not aid_tag:
                continue
            aid = re.search(r"/answer/(\d+)", aid_tag["href"])
            if not aid:
                continue
            aid = aid.group(1)
            if aid not in seen_ids:
                seen_ids.add(aid)
                new_in_last += 1
                all_cards.append(c)

        print(f"[Zhihu] 关键词 “{keyword}” 第 {page} 页抓到 {new_in_last} 条新回答，累计 {len(seen_ids)} 条")

        if new_in_last == 0:
            print("[Zhihu] 已无更多新内容，提前终止翻页")
            break
        time.sleep(random.uniform(3, 5))

    return all_cards


def zhihu_parse_answer_card(card, keyword: str) -> dict:
    rec = {
        "关键词": keyword,
        "回答ID": "",
        "问题ID": "",
        "问题标题": "",
        "回答链接": "",
        "回答正文": "",
        "回答创建时间": "",
        "回答更新时间": "",
        "点赞数": 0,
        "评论数": 0,
        "作者ID": "",
        "作者昵称": "",
        "作者性别": "未知",
        "作者头像": "",
        "作者主页": "",
        "作者粉丝数": 0,
        "评论": [],
        "抓取时间": datetime.now().isoformat(),
    }

    q_tag = card.select_one('a[href*="/question/"]')
    if q_tag:
        qid = re.search(r"/question/(\d+)", q_tag["href"])
        if qid:
            rec["问题ID"] = qid.group(1)
            rec["问题标题"] = q_tag.get_text(strip=True)

    a_tag = card.select_one('a[href*="/answer/"]')
    if a_tag:
        aid = re.search(r"/answer/(\d+)", a_tag["href"])
        if aid:
            rec["回答ID"] = aid.group(1)
            rec["回答链接"] = "https://www.zhihu.com" + a_tag["href"].split("?")[0]

    rich = card.select_one("span.RichText")
    if rich:
        rec["回答正文"] = rich.get_text(" ", strip=True)

    vote = card.select_one("button.Button.VoteButton--up")
    if vote:
        n = re.search(r"\d+", vote.text)
        rec["点赞数"] = int(n.group()) if n else 0
    cmt = card.select_one("button.ContentItem-action")
    if cmt:
        n = re.search(r"\d+", cmt.text)
        rec["评论数"] = int(n.group()) if n else 0

    author = card.select_one("a.UserLink-link")
    if author:
        rec["作者昵称"] = author.get_text(strip=True)
        token = author["href"].split("/")[-1]
        rec["作者主页"] = f"https://www.zhihu.com/people/{token}"
        rec["作者ID"] = token
    avatar = card.select_one("img.Avatar")
    if avatar:
        rec["作者头像"] = avatar["src"]

    return rec


def zhihu_save_single_excel(keyword: str, rows: List[dict], output_dir: str = ZHIHU_OUTPUT_DIR) -> Optional[str]:
    if not rows:
        return None
    try:
        os.makedirs(output_dir, exist_ok=True)
        safe_kw = _sanitize_filename(keyword)
        filepath = os.path.join(output_dir, f"zhihu_{safe_kw}.xlsx")
        df = pd.DataFrame(rows)
        df.to_excel(filepath, index=False, engine="openpyxl")
        print(f"[Zhihu] 单关键词Excel已保存 -> {filepath}")
        return filepath
    except Exception as e:
        print(f"[Zhihu] 保存关键词 {keyword} 的Excel时出错: {e}")
        return None


def zhihu_save_combined_excel(all_data: List[tuple], output_dir: str = ZHIHU_OUTPUT_DIR) -> Optional[str]:
    """
    all_data: List of (keyword, rows)
    """
    if not all_data:
        print("[Zhihu] 无数据可导出整合文件。")
        return None
    try:
        os.makedirs(output_dir, exist_ok=True)
        combined_rows = []
        for keyword, rows in all_data:
            if not rows:
                continue
            for r in rows:
                rec = dict(r)
                rec["关键词"] = keyword  # ensure keyword present
                combined_rows.append(rec)

        if not combined_rows:
            print("[Zhihu] 无数据可导出整合文件。")
            return None

        filepath = os.path.join(output_dir, "zhihu_all_combined.xlsx")
        pd.DataFrame(combined_rows).to_excel(filepath, index=False, engine="openpyxl")
        print(f"[Zhihu] 整合Excel已导出 -> {filepath}")
        return filepath
    except Exception as e:
        print(f"[Zhihu] 导出整合Excel时出错: {e}")
        return None


def zhihu_run_keywords(keywords: List[str], headless: bool = False) -> int:
    """
    批量运行知乎关键词爬取，返回成功处理的关键词数量。
    """
    if not keywords:
        print("[Zhihu] 关键词列表为空，跳过知乎爬取。")
        return 0

    zhihu_ensure_dir()
    done = zhihu_load_checkpoint()
    todo = [k for k in keywords if k not in done]
    if not todo:
        print("[Zhihu] 所有关键词已抓完，如需重抓请删除 zhihu/checkpoint.txt")
        return 0

    zhihu_all_data: List[tuple] = []

    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    if headless:
        options.add_argument("--headless=new")

    driver = webdriver.Chrome(options=options)

    try:
        if not zhihu_load_driver_cookies(driver):
            zhihu_login_by_selenium(driver)

        success_count = 0
        total = len(todo)
        for idx, kw in enumerate(todo, 1):
            print(f"\n[Zhihu] 进度：{idx}/{total} | 关键词：{kw}")
            cards = zhihu_search_by_keyword_dom(kw, driver)
            rows = []
            for c in cards:
                row = zhihu_parse_answer_card(c, kw)
                zhihu_append_one_row(row)
                rows.append(row)
            if rows:
                zhihu_save_single_excel(kw, rows, ZHIHU_OUTPUT_DIR)
                zhihu_all_data.append((kw, rows))
            zhihu_save_checkpoint(kw)
            success_count += 1
            print(f"[Zhihu] 关键词 “{kw}” 完成，已追加 {len(cards)} 条")
            if idx < total:
                time.sleep(random.uniform(2.0, 4.0))
        zhihu_save_combined_excel(zhihu_all_data, ZHIHU_OUTPUT_DIR)
        return success_count
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        zhihu_close_csv()


def _sanitize_filename(name: str) -> str:
    """移除不安全的文件名字符并压缩空白。"""
    safe = re.sub(r"[\\/:*?\"<>|]", "_", name)
    safe = re.sub(r"\s+", "_", safe).strip("_")
    return safe or "keyword"

# ======================
# 豆瓣爬虫内联实现（由 douban.py 嵌入精简版）
# ======================
DOUBAN_TARGET_COUNT = 200
DOUBAN_HEADLESS = False
DOUBAN_SCROLL_PAUSE_RANGE = (1.0, 2.2)
DOUBAN_ELEMENT_WAIT_SECONDS = 15
DOUBAN_PAGE_MAX_FAIL = 6
DOUBAN_MANUAL_COOKIE_STRING: Optional[str] = None
DOUBAN_QR_LOGIN_WAIT_SECONDS = 480  # 扫码等待时长延长到8分钟
DOUBAN_COOKIE_FILE = os.path.join(DEFAULT_OUTPUT_DIR, "douban_login_cookies.json")
DOUBAN_RESULT_SELECTORS = [
    "ul.search-result li",  # 经典搜索结果列表
    "li.result",  # 备用
    "li.search-item",  # 备用
    "li.item",  # 旧版
    "div.result",  # 另一种布局
    "div.result-item",  # 备用
    "div.article",  # 备用
]


def douban_human_pause(a=0.8, b=1.6):
    time.sleep(random.uniform(a, b))


def douban_rand_in_range(r):
    return random.uniform(r[0], r[1])


def douban_scroll_for_results(driver: webdriver.Chrome, steps: int = 4):
    """向下滚动多次，触发懒加载。"""
    try:
        for _ in range(max(1, steps)):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            douban_human_pause(*DOUBAN_SCROLL_PAUSE_RANGE)
    except Exception:
        pass


def douban_wait_for_results(driver: webdriver.Chrome, timeout: int = DOUBAN_ELEMENT_WAIT_SECONDS) -> Optional[str]:
    """
    等待搜索结果列表出现，返回命中的selector；若未出现返回None。
    """
    for sel in DOUBAN_RESULT_SELECTORS:
        try:
            WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
            return sel
        except Exception:
            continue
    return None


def douban_parse_number_from_text(text: str, keywords: List[str] = None) -> str:
    if not text:
        return "无数据"
    s = text.strip()
    if re.match(r"^\d+$", s):
        return s
    unit_match = re.search(r"(\d+(?:\.\d+)?)\s*([万Ww])", s)
    if unit_match:
        try:
            num = float(unit_match.group(1)) * 10000
            return str(int(num))
        except Exception:
            pass
    comma_match = re.search(r"(\d{1,3}(?:,\d{3})+)", s)
    if comma_match:
        try:
            num_str = comma_match.group(1).replace(",", "")
            num = float(num_str)
            return str(int(num))
        except Exception:
            pass
    all_numbers = re.findall(r"\d+", s)
    if all_numbers:
        longest_num = max(all_numbers, key=len)
        try:
            num = float(longest_num)
            return str(int(num))
        except Exception:
            pass
    if keywords:
        for kw in keywords:
            s = s.replace(kw, "")
    s = s.replace("+", "").replace(",", "").strip()
    if not s:
        return "无数据"
    if re.match(r"^\d+$", s):
        return s
    m = re.match(r"^([0-9]+(?:\.[0-9]+)?)\s*[万Ww]$", s)
    if m:
        try:
            val = float(m.group(1)) * 10000
            return str(int(val))
        except Exception:
            return "无数据"
    m3 = re.match(r"^([0-9]+(?:\.[0-9]+)?)$", s)
    if m3:
        try:
            f = float(m3.group(1))
            return str(int(f))
        except Exception:
            return "无数据"
    m4 = re.search(r"([0-9]+(?:\.[0-9]+)?)([万Ww]?)", s)
    if m4:
        num = float(m4.group(1))
        if m4.group(2) in ["万", "W", "w"]:
            num *= 10000
        return str(int(num))
    return "无数据"


def douban_parse_reply_to_number(text: str) -> str:
    return douban_parse_number_from_text(text, keywords=["回复"])


def douban_parse_view_to_number(text: str) -> str:
    return douban_parse_number_from_text(text, keywords=["浏览"])


def douban_filter_post_content(text: str) -> str:
    if not text:
        return ""
    s = text.strip()
    image_patterns = [r"图片\d+", r"查看图片", r"查看原图", r"图片来自", r"图\d+"]
    ad_patterns = [r"广告", r"推广", r"点击查看", r"更多详情"]
    image_count = sum(1 for p in image_patterns if re.search(p, s, re.IGNORECASE))
    ad_count = sum(1 for p in ad_patterns if re.search(p, s, re.IGNORECASE))
    if image_count > 0 and len(s) < 50:
        return ""
    if ad_count > 0 and len(s) < 30:
        return ""
    for pattern in image_patterns + ad_patterns:
        s = re.sub(pattern, "", s, flags=re.IGNORECASE)
    return s.strip()


def douban_clean_text(text: str) -> str:
    if not text:
        return ""
    s = re.sub(r"\s+", " ", text).strip()
    patterns_to_remove = [
        r"回复\s*\d+(?:\.\d+)?[万Ww]?\+?",
        r"浏览\s*\d+(?:\.\d+)?[万Ww]?\+?",
        r"\d+(?:\.\d+)?[万Ww]?\s*回复",
        r"\d+(?:\.\d+)?[万Ww]?\s*浏览",
    ]
    for pattern in patterns_to_remove:
        s = re.sub(pattern, "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def douban_sanitize_keyword(raw: str) -> str:
    """
    去掉关键词中括号及其内容（含中英文括号），再做简单空白折叠。
    """
    if not raw:
        return ""
    cleaned = re.sub(r"\s*[\(\（][^\)\）]*[\)\）]\s*", " ", str(raw))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or str(raw).strip()


def douban_build_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    if DOUBAN_HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1440,900")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(45)
    driver.set_script_timeout(45)
    return driver


def douban_inject_manual_cookie_if_any(driver: webdriver.Chrome):
    if not DOUBAN_MANUAL_COOKIE_STRING:
        return
    try:
        driver.get("https://www.douban.com/")
        douban_human_pause(1.2, 2.0)
        parts = [p.strip() for p in DOUBAN_MANUAL_COOKIE_STRING.split(";") if "=" in p]
        for p in parts:
            name, value = p.split("=", 1)
            driver.add_cookie({"name": name.strip(), "value": value.strip(), "domain": ".douban.com"})
        driver.get("https://www.douban.com/")
        douban_human_pause(1.0, 1.6)
    except Exception:
        pass


def douban_save_cookies(driver: webdriver.Chrome, path: str = DOUBAN_COOKIE_FILE):
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(driver.get_cookies(), f, ensure_ascii=False, indent=2)
        print(f"[Douban] 登录状态已保存 -> {path}")
    except Exception as e:
        print(f"[Douban] 保存 cookies 失败: {e}")


def douban_load_cookies(driver: webdriver.Chrome, path: str = DOUBAN_COOKIE_FILE) -> bool:
    try:
        if not os.path.exists(path):
            return False
        driver.get("https://www.douban.com/")
        douban_human_pause(0.8, 1.5)
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
        for c in cookies:
            try:
                driver.add_cookie(c)
            except Exception:
                continue
        driver.get("https://www.douban.com/")
        douban_human_pause(1.0, 1.6)
        print(f"[Douban] 已加载本地 cookies -> {path}")
        return True
    except Exception as e:
        print(f"[Douban] 读取 cookies 失败: {e}")
        return False


def douban_wait_for_manual_login(driver: webdriver.Chrome, max_seconds: int) -> bool:
    deadline = time.time() + max_seconds
    last_hint = 0.0
    while time.time() < deadline:
        try:
            cookies = driver.get_cookies()
            login_cookies = [c for c in cookies if (c.get("name") or "").lower() in ("bid", "dbcl2")]
            if login_cookies:
                print(f"检测到登录Cookie: {login_cookies[0].get('name')} = {login_cookies[0].get('value')[:10]}...")
                return True
        except Exception as e:
            print(f"Cookie检查出错: {e}")
        now = time.time()
        if now - last_hint > 5:
            last_hint = now
            remaining = int(deadline - now)
            print(f"等待扫码登录中...（剩余约 {remaining}s）| 当前URL: {driver.current_url[:50]}...")
        time.sleep(5)
    return False


def douban_is_logged_in(driver: webdriver.Chrome) -> bool:
    try:
        for c in driver.get_cookies():
            name = (c.get("name") or "").lower()
            if name in ("bid", "dbcl2"):
                return True
    except Exception:
        pass
    return False


def douban_ensure_logged_in(driver: webdriver.Chrome, max_seconds: int = DOUBAN_QR_LOGIN_WAIT_SECONDS) -> bool:
    """
    登录校验：只有在明确检测到已登录后才继续。
    1) 先尝试加载本地 cookies；若成功直接返回。
    2) 未登录时只停留在首页，等待用户手动登录（不自动跳转/点击），完成后继续。
    3) 登录成功立即保存 cookies。
    """
    # 尝试加载本地 cookie（需要先进入主站域以便写入）
    try:
        driver.get("https://www.douban.com/")
        WebDriverWait(driver, DOUBAN_ELEMENT_WAIT_SECONDS).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except Exception:
        pass

    if douban_load_cookies(driver) and douban_is_logged_in(driver):
        print("[Douban] 已通过本地 cookies 自动登录。")
        douban_save_cookies(driver)
        return True

    # 再次快速检测（避免重复走扫码）
    if douban_is_logged_in(driver):
        print("[Douban] 已检测到登录态，直接开始采集。")
        douban_save_cookies(driver)
        return True

    if DOUBAN_HEADLESS and not DOUBAN_MANUAL_COOKIE_STRING:
        print("[Douban] 无头模式且未配置 DOUBAN_MANUAL_COOKIE_STRING，无法完成登录校验。")
        return False

    # 停留在首页，由用户自行登录（不自动跳转/点击）
    if not DOUBAN_MANUAL_COOKIE_STRING and not DOUBAN_HEADLESS:
        print(f"[Douban] 请在当前浏览器页面自行登录，系统将等待最多 {max_seconds} 秒...")
        print("[Douban] 登录成功后保持页面打开以便检测。")
        ok_login = douban_wait_for_manual_login(driver, max_seconds)
    else:
        ok_login = False

    if douban_is_logged_in(driver):
        print("[Douban] 登录验证通过，开始采集。")
        douban_save_cookies(driver)
        return True

    if ok_login:
        print("[Douban] 检测到扫码流程完成，但未识别到登录 Cookie，请重试或配置 DOUBAN_MANUAL_COOKIE_STRING。")
    else:
        print("[Douban] 未检测到登录，程序终止。")
    return False


def douban_run_setup() -> None:
    """
    预登录模式：仅打开豆瓣登录页并保存登录态，供后续爬虫复用。
    """
    print("[Douban] 进入预登录模式，将打开豆瓣登录页以便完成登录...")
    original_headless = DOUBAN_HEADLESS
    globals()["DOUBAN_HEADLESS"] = False  # 预登录必须有头，便于扫码/输入
    try:
        driver = douban_build_driver()
    except Exception as e:
        print(f"[Douban] 浏览器初始化失败，无法进行预登录: {e}")
        globals()["DOUBAN_HEADLESS"] = original_headless
        return

    try:
        login_url = "https://accounts.douban.com/passport/login?redir=https://www.douban.com/"
        try:
            driver.get(login_url)
            print("[Douban] 已打开登录页，请使用手机扫码或输入账号密码完成登录。")
        except Exception:
            print("[Douban] 打开登录页失败，将尝试回退到首页。")
            try:
                driver.get("https://www.douban.com/")
            except Exception:
                pass
        douban_inject_manual_cookie_if_any(driver)
        print("[Douban] 请在浏览器中完成登录，完成后按回车以保存登录状态（窗口不会自动关闭）。")
        try:
            input(">>> 登录完成后按回车继续...")
        except Exception:
            pass
        # 用户确认后再校验登录态并保存
        if douban_is_logged_in(driver):
            douban_save_cookies(driver)
            print(f"[Douban] 登录成功，登录态已写入 {DOUBAN_COOKIE_FILE}，后续爬虫将自动复用。")
        else:
            # 兜底再尝试一次标准登录校验（含等待）
            ok = douban_ensure_logged_in(driver, DOUBAN_QR_LOGIN_WAIT_SECONDS)
            if ok and douban_is_logged_in(driver):
                douban_save_cookies(driver)
                print(f"[Douban] 登录成功，登录态已写入 {DOUBAN_COOKIE_FILE}，后续爬虫将自动复用。")
            else:
                print("[Douban] 未能完成预登录，未保存登录状态。")
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        globals()["DOUBAN_HEADLESS"] = original_headless


def douban_goto_search_result_direct(driver: webdriver.Chrome, keyword: str):
    encoded_q = quote(keyword)
    url = f"https://www.douban.com/group/search?cat=1013&q={encoded_q}"
    driver.get(url)
    WebDriverWait(driver, DOUBAN_ELEMENT_WAIT_SECONDS).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    douban_human_pause(1.0, 1.6)


def douban_is_result_page(driver: webdriver.Chrome) -> bool:
    u = driver.current_url
    return ("douban.com/group/search" in u) or ("search" in u) or ("q=" in u)


def douban_extract_interaction_numbers(post_element) -> Dict[str, str]:
    result = {"Reply Count": "无数据", "View Count": "无数据"}
    try:
        reply_elements = post_element.find_elements(By.CSS_SELECTOR, "span[class*='count'], span[class*='reply']")
        for elem in reply_elements:
            try:
                text = (elem.text or "").strip()
                if "回复" in text:
                    result["Reply Count"] = douban_parse_reply_to_number(text)
                    break
            except Exception:
                continue
        view_elements = post_element.find_elements(By.CSS_SELECTOR, "span[class*='count'], span[class*='view']")
        for elem in view_elements:
            try:
                text = (elem.text or "").strip()
                if "浏览" in text:
                    result["View Count"] = douban_parse_view_to_number(text)
                    break
            except Exception:
                continue
        full_text = post_element.text.strip()
        if result["Reply Count"] == "无数据":
            reply_match = re.search(r"回复\s*(\d+(?:\.\d+)?[万Ww]?)", full_text)
            if reply_match:
                result["Reply Count"] = douban_parse_reply_to_number(reply_match.group(0))
        if result["View Count"] == "无数据":
            view_match = re.search(r"浏览\s*(\d+(?:\.\d+)?[万Ww]?)", full_text)
            if view_match:
                result["View Count"] = douban_parse_view_to_number(view_match.group(0))
    except Exception:
        pass
    return result


def douban_collect_posts_from_dom(driver: webdriver.Chrome) -> List[Dict[str, str]]:
    results: List[Dict[str, str]] = []
    try:
        items = []
        seen_ids = set()
        for sel in DOUBAN_RESULT_SELECTORS:
            found = driver.find_elements(By.CSS_SELECTOR, sel)
            for elem in found:
                if elem.id not in seen_ids:
                    seen_ids.add(elem.id)
                    items.append(elem)
        # 兜底：直接找链接节点
        link_nodes = driver.find_elements(By.CSS_SELECTOR, "a[href*='douban.com/group/topic']")
        for ln in link_nodes:
            try:
                container = ln.find_element(By.XPATH, "./ancestor::li | ./ancestor::div[contains(@class,'result') or contains(@class,'article')]")
                if container.id not in seen_ids:
                    seen_ids.add(container.id)
                    items.append(container)
            except Exception:
                continue
        # 旧版表格布局（如示例截图）
        table_rows = driver.find_elements(By.CSS_SELECTOR, "table.olt tr")
        for tr in table_rows[1:]:  # 跳过表头
            try:
                tds = tr.find_elements(By.TAG_NAME, "td")
                if len(tds) < 3:
                    continue
                title_links = tds[0].find_elements(By.CSS_SELECTOR, "a[href*='douban.com/group/topic']")
                if not title_links:
                    continue
                title_link = title_links[0]
                title = (title_link.text or "").strip()
                link = title_link.get_attribute("href") or ""
                time_str = (tds[1].text or "").strip() if len(tds) > 1 else ""
                reply_text = (tds[2].text or "").strip() if len(tds) > 2 else ""
                results.append(
                    {
                        "Title": title,
                        "Author": "N/A",
                        "Time": time_str or "无数据",
                        "Text Content": "",
                        "Reply Count": douban_parse_reply_to_number(reply_text) if reply_text else "无数据",
                        "View Count": "无数据",
                        "Post Link": link or "无数据",
                    }
                )
            except Exception:
                continue

        print(f"找到 {len(items)} 个潜在帖子项")
        for item in items:
            try:
                title_elems = item.find_elements(By.CSS_SELECTOR, "h3 a, .title a")
                if not title_elems:
                    continue
                title = title_elems[0].text.strip() if title_elems else ""
                author = ""
                author_elems = item.find_elements(By.CSS_SELECTOR, ".user a, .author a, span.author")
                if author_elems:
                    author = author_elems[0].text.strip()
                time_str = ""
                time_elems = item.find_elements(By.CSS_SELECTOR, ".created, .time, span.time")
                if time_elems:
                    time_str = time_elems[0].text.strip()
                content = ""
                content_nodes = item.find_elements(By.CSS_SELECTOR, ".topic-content p, .content p, div[dir='auto']")
                if content_nodes:
                    main_text = content_nodes[0].text.strip()
                    content = douban_clean_text(douban_filter_post_content(main_text))[:200]
                if not content:
                    text_parts = []
                    for node in content_nodes:
                        txt = node.text.strip()
                        if txt and len(txt) > 5:
                            filtered = douban_filter_post_content(txt)
                            if filtered:
                                text_parts.append(filtered)
                    if text_parts:
                        content = douban_clean_text(" ".join(text_parts))[:200]
                interaction_data = douban_extract_interaction_numbers(item)
                reply_text = interaction_data.get("Reply Count", "无数据")
                view_text = interaction_data.get("View Count", "无数据")
                link = ""
                if title_elems:
                    link = title_elems[0].get_attribute("href") or ""
                if title or content:
                    results.append(
                        {
                            "Title": title or "",
                            "Author": author or "N/A",
                            "Time": time_str or "无数据",
                            "Text Content": content or "",
                            "Reply Count": reply_text if (reply_text and reply_text.strip()) else "0",
                            "View Count": view_text if (view_text and view_text.strip()) else "0",
                            "Post Link": link or "无数据",
                        }
                    )
            except Exception as e:
                print(f"处理帖子时出错: {str(e)}")
                continue
    except Exception as e:
        print(f"提取帖子时出错: {str(e)}")
    return results


def douban_crawl_keyword(driver: webdriver.Chrome, keyword: str, script_dir: str) -> Optional[List[Dict[str, str]]]:
    try:
        print(f"\n{'='*60}")
        print(f"开始爬取关键词：{keyword}")
        print(f"{'='*60}\n")
        douban_goto_search_result_direct(driver, keyword)
        douban_wait_for_results(driver)
        douban_scroll_for_results(driver, steps=4)
        current_url = driver.current_url
        print(f"当前页面URL: {current_url}")
        if not douban_is_result_page(driver):
            print("警告：可能未成功到达搜索结果页面")
            return None
        douban_human_pause(2.0, 3.0)
        collected = douban_collect_posts_from_dom(driver)
        for row in collected:
            row.setdefault("Title", "")
            row.setdefault("Author", "N/A")
            row.setdefault("Time", "无数据")
            row.setdefault("Text Content", "")
            row.setdefault("Reply Count", "无数据")
            row.setdefault("View Count", "无数据")
            row.setdefault("Post Link", "无数据")
        print(f"\n关键词 '{keyword}' 爬取完成！共收集 {len(collected)} 条数据\n")
        return collected
    except Exception as e:
        print(f"\n爬取关键词 '{keyword}' 时发生异常：{str(e)}")
        traceback.print_exc()
        return None


def douban_save_single_excel(keyword: str, data: List[Dict[str, str]], out_dir: str):
    """单关键词即时导出，便于长跑中随时查看结果。"""
    if not data:
        return None
    os.makedirs(out_dir, exist_ok=True)
    safe_kw = _sanitize_filename(keyword) or "keyword"
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"douban_{safe_kw}_{timestamp}.xlsx"
    filepath = os.path.join(out_dir, filename)
    df = pd.DataFrame(
        data,
        columns=["Title", "Author", "Time", "Text Content", "Reply Count", "View Count", "Post Link"],
    )
    df.insert(0, "Key Word", keyword)
    df.to_excel(filepath, index=False, sheet_name="Sheet1")
    print(f"[Douban] 已导出单关键词文件：{filepath}")
    return filepath


def douban_save_combined_excel(all_data: List[Tuple[str, List[Dict[str, str]]]], script_dir: str):
    try:
        date_str = dt.datetime.now().strftime("%Y%m%d")
        filename = f"豆瓣搜索结果_整合_{date_str}.xlsx"
        filepath = os.path.join(script_dir, filename)
        combined_rows = []
        valid_data = [(k, d) for k, d in all_data if d]
        for keyword, data_list in valid_data:
            for row in data_list:
                combined_row = {"Key Word": keyword}
                combined_row.update(row)
                combined_rows.append(combined_row)
        if not combined_rows:
            print("警告：没有数据可导出！")
            return False
        df = pd.DataFrame(
            combined_rows,
            columns=["Key Word", "Title", "Author", "Time", "Text Content", "Reply Count", "View Count", "Post Link"],
        )
        for col in ["Reply Count", "View Count"]:
            df[col] = df[col].apply(lambda x: str(x) if x is not None and x != "" else ("无数据" if x == "" else ""))
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]
            for row_idx in range(2, len(df) + 2):
                for col_idx in [6, 7]:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = df.iloc[row_idx - 2, col_idx - 1]
                    cell_value_str = str(cell_value) if cell_value is not None and cell_value != "" else ""
                    cell.number_format = "@"
                    cell.value = cell_value_str
        total_rows = sum(len(data_list) for _, data_list in valid_data)
        print(f"\n{'='*60}")
        print(f"整合Excel文件已导出：{filepath}")
        print(f"共包含 {len(valid_data)} 个关键词，总计 {total_rows} 条数据")
        print(f"{'='*60}\n")
        return True
    except Exception as e:
        print(f"\n导出整合Excel文件时出错：{str(e)}")
        traceback.print_exc()
        return False

# ======================
# 微博爬虫内联实现（从 weibo.py 嵌入）
# ======================
WEIBO_TARGET_COUNT = 200
WEIBO_HEADLESS = False
WEIBO_SCROLL_PAUSE_RANGE = (1.0, 2.2)
WEIBO_ELEMENT_WAIT_SECONDS = 15
WEIBO_PAGE_MAX_FAIL = 6
WEIBO_MANUAL_COOKIE_STRING: Optional[str] = None
WEIBO_QR_LOGIN_WAIT_SECONDS = 120


def weibo_human_pause(a=0.8, b=1.6):
    time.sleep(random.uniform(a, b))


def weibo_rand_in_range(r):
    return random.uniform(r[0], r[1])


def weibo_parse_number_from_text(text: str, keywords: List[str] = None) -> str:
    if not text:
        return "无数据"
    s = text.strip()
    if re.match(r"^\d+$", s):
        return s
    unit_match = re.search(r"(\d+(?:\.\d+)?)\s*([万Ww])", s)
    if unit_match:
        try:
            num = float(unit_match.group(1)) * 10000
            return str(int(num))
        except Exception:
            pass
    comma_match = re.search(r"(\d{1,3}(?:,\d{3})+)", s)
    if comma_match:
        try:
            num = float(comma_match.group(1).replace(",", ""))
            return str(int(num))
        except Exception:
            pass
    all_numbers = re.findall(r"\d+", s)
    if all_numbers:
        try:
            return str(int(float(max(all_numbers, key=len))))
        except Exception:
            pass
    if keywords:
        for kw in keywords:
            s = s.replace(kw, "")
    s = s.replace("+", "").replace(",", "").strip()
    if not s:
        return "无数据"
    if re.match(r"^\d+$", s):
        return s
    m = re.match(r"^([0-9]+(?:\.[0-9]+)?)\s*[万Ww]$", s)
    if m:
        try:
            return str(int(float(m.group(1)) * 10000))
        except Exception:
            return "无数据"
    m3 = re.match(r"^([0-9]+(?:\.[0-9]+)?)$", s)
    if m3:
        try:
            return str(int(float(m3.group(1))))
        except Exception:
            return "无数据"
    m4 = re.search(r"([0-9]+(?:\.[0-9]+)?)([万Ww]?)", s)
    if m4:
        num = float(m4.group(1))
        if m4.group(2) in ["万", "W", "w"]:
            num *= 10000
        return str(int(num))
    return "无数据"


def weibo_filter_post_content(text: str) -> str:
    if not text:
        return ""
    s = text.strip()
    image_ocr_patterns = [r"图片\d+", r"查看图片", r"查看原图", r"图片来自", r"图\d+"]
    app_patterns = [r"投票", r"投票器", r"小程序", r"点击参与", r"参与投票", r"查看详情", r"立即参与"]
    image_count = sum(1 for p in image_ocr_patterns if re.search(p, s, re.IGNORECASE))
    app_count = sum(1 for p in app_patterns if re.search(p, s, re.IGNORECASE))
    if image_count > 0 and len(s) < 50:
        return ""
    if app_count > 0 and len(s) < 30:
        return ""
    for p in image_ocr_patterns + app_patterns:
        s = re.sub(p, "", s, flags=re.IGNORECASE)
    return s.strip()


def weibo_clean_text(text: str) -> str:
    if not text:
        return ""
    s = re.sub(r"\s+", " ", text).strip()
    patterns_to_remove = [
        r"转发\s*\d+(?:\.\d+)?[万Ww]?\+?",
        r"评论\s*\d+(?:\.\d+)?[万Ww]?\+?",
        r"赞\s*\d+(?:\.\d+)?[万Ww]?\+?",
        r"\d+(?:\.\d+)?[万Ww]?\s*转发",
        r"\d+(?:\.\d+)?[万Ww]?\s*评论",
        r"\d+(?:\.\d+)?[万Ww]?\s*赞",
    ]
    for p in patterns_to_remove:
        s = re.sub(p, "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def weibo_build_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    if WEIBO_HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1440,900")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(45)
    driver.set_script_timeout(45)
    return driver


def weibo_inject_manual_cookie_if_any(driver: webdriver.Chrome):
    if not WEIBO_MANUAL_COOKIE_STRING:
        return
    try:
        driver.get("https://www.weibo.com/")
        weibo_human_pause(1.2, 2.0)
        parts = [p.strip() for p in WEIBO_MANUAL_COOKIE_STRING.split(";") if "=" in p]
        for p in parts:
            name, value = p.split("=", 1)
            driver.add_cookie({"name": name.strip(), "value": value.strip(), "domain": ".weibo.com"})
        driver.get("https://www.weibo.com/")
        weibo_human_pause(1.0, 1.6)
    except Exception:
        pass


def weibo_wait_for_manual_login(driver: webdriver.Chrome, max_seconds: int) -> bool:
    deadline = time.time() + max_seconds
    last_hint = 0.0
    last_refresh = 0.0
    while time.time() < deadline:
        try:
            for c in driver.get_cookies():
                name = (c.get("name") or "").upper()
                if name in ("SUB", "SUBP"):
                    return True
        except Exception:
            pass
        now = time.time()
        if now - last_refresh > 10:
            last_refresh = now
            try:
                driver.get("https://www.weibo.com/")
            except Exception:
                pass
            time.sleep(1.0)
        if now - last_hint > 5:
            last_hint = now
            print("等待扫码登录中...（剩余约 %ds）" % int(deadline - now))
        time.sleep(1.5)
    return False


def weibo_is_logged_in(driver: webdriver.Chrome) -> bool:
    try:
        for c in driver.get_cookies():
            if (c.get("name") or "").upper() in ("SUB", "SUBP"):
                return True
    except Exception:
        pass
    return False


def weibo_try_homepage_search(driver: webdriver.Chrome, keyword: str) -> bool:
    try:
        driver.get("https://www.weibo.com/")
        WebDriverWait(driver, WEIBO_ELEMENT_WAIT_SECONDS).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        weibo_human_pause()
        try:
            close_btns = driver.find_elements(By.CSS_SELECTOR, '[aria-label*="关闭"], [title*="关闭"], .woo-dialog-actions .woo-button-main')
            for b in close_btns[:2]:
                try:
                    b.click()
                    weibo_human_pause(0.4, 0.8)
                except Exception:
                    pass
        except Exception:
            pass
        candidates = [
            'input[placeholder*="搜索"]',
            'input[type="search"]',
            '[data-testid="search-input"] input',
            'input.suggest-input',
        ]
        search_input = None
        for css in candidates:
            elems = driver.find_elements(By.CSS_SELECTOR, css)
            if elems:
                search_input = elems[0]
                break
        if not search_input:
            return False
        ActionChains(driver).move_to_element(search_input).pause(0.3).click().perform()
        weibo_human_pause(0.2, 0.5)
        search_input.clear()
        for ch in keyword:
            search_input.send_keys(ch)
            time.sleep(random.uniform(0.05, 0.12))
        search_input.submit()
        weibo_human_pause(1.2, 1.8)
        WebDriverWait(driver, WEIBO_ELEMENT_WAIT_SECONDS).until(EC.url_contains("weibo.com"))
        return True
    except Exception:
        return False


def weibo_goto_search_result_direct(driver: webdriver.Chrome, keyword: str):
    url = f"https://s.weibo.com/weibo?q={keyword}"
    driver.get(url)
    WebDriverWait(driver, WEIBO_ELEMENT_WAIT_SECONDS).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    weibo_human_pause(1.0, 1.6)


def weibo_is_result_page(driver: webdriver.Chrome) -> bool:
    u = driver.current_url
    return ("s.weibo.com/weibo" in u) or ("search" in u) or ("weibo.com" in u and "q=" in u)


def weibo_extract_interaction_numbers(post_element) -> Dict[str, str]:
    result = {"Repost Count": "无数据", "Comment Count": "无数据", "Like Count": "无数据"}
    try:
        forward_elements = post_element.find_elements(By.CSS_SELECTOR, '[action-type="feed_list_forward"], [action-type="fl_forward"]')
        for elem in forward_elements:
            text = (elem.text or "").strip()
            if text:
                nums = re.findall(r"\d+", text)
                if nums:
                    result["Repost Count"] = nums[0]
                    break
        comment_elements = post_element.find_elements(By.CSS_SELECTOR, '[action-type="feed_list_comment"], [action-type="fl_comment"]')
        for elem in comment_elements:
            text = (elem.text or "").strip()
            if text:
                nums = re.findall(r"\d+", text)
                if nums:
                    result["Comment Count"] = nums[0]
                    break
        like_elements = post_element.find_elements(By.CSS_SELECTOR, '[action-type="feed_list_like"], [action-type="fl_like"]')
        for elem in like_elements:
            like_count_elem = elem.find_elements(By.CSS_SELECTOR, ".woo-like-count, [class*='like-count']")
            text = ""
            if like_count_elem:
                text = (like_count_elem[0].text or "").strip()
            else:
                text = (elem.text or "").strip()
            if text:
                nums = re.findall(r"\d+", text)
                if nums:
                    result["Like Count"] = nums[0]
                    break
        if "无数据" in result.values():
            card_act = post_element.find_elements(By.CSS_SELECTOR, ".card-act")
            if card_act:
                for link in card_act[0].find_elements(By.CSS_SELECTOR, "a"):
                    action_type = (link.get_attribute("action-type") or "").strip()
                    text = (link.text or "").strip()
                    if not action_type or not text:
                        continue
                    nums = re.findall(r"\d+", text)
                    if not nums:
                        continue
                    num = nums[0]
                    if action_type in ("fl_forward", "feed_list_forward") and result["Repost Count"] == "无数据":
                        result["Repost Count"] = num
                    elif action_type in ("fl_comment", "feed_list_comment") and result["Comment Count"] == "无数据":
                        result["Comment Count"] = num
                    elif action_type in ("fl_like", "feed_list_like") and result["Like Count"] == "无数据":
                        result["Like Count"] = num
    except Exception:
        pass
    return result


def weibo_expand_folded_posts(driver: webdriver.Chrome):
    try:
        expanded_count = 0
        max_attempts = 3
        for _ in range(max_attempts):
            found_any = False
            selectors = ['a[action-type="fl_unfold"]', 'a[node-type="feed_list_content_full"]']
            for selector in selectors:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    try:
                        href = elem.get_attribute("href") or ""
                        if re.search(r"/status/\d+", href):
                            continue
                        if elem.is_displayed() and elem.is_enabled():
                            driver.execute_script("arguments[0].scrollIntoView({block:'center', behavior:'smooth'});", elem)
                            weibo_human_pause(0.2, 0.4)
                            elem.click()
                            found_any = True
                            expanded_count += 1
                            weibo_human_pause(0.5, 0.8)
                    except Exception:
                        continue
            xpath_selectors = ["//a[contains(text(), '展开全文')]"]
            for xpath in xpath_selectors:
                elements = driver.find_elements(By.XPATH, xpath)
                for elem in elements:
                    try:
                        href = elem.get_attribute("href") or ""
                        if re.search(r"/status/\d+", href):
                            continue
                        if elem.is_displayed() and elem.is_enabled():
                            driver.execute_script("arguments[0].scrollIntoView({block:'center', behavior:'smooth'});", elem)
                            weibo_human_pause(0.2, 0.4)
                            elem.click()
                            found_any = True
                            expanded_count += 1
                            weibo_human_pause(0.5, 0.8)
                    except Exception:
                        continue
            if not found_any:
                break
            weibo_human_pause(0.3, 0.5)
        if expanded_count > 0:
            print(f"批量展开：已展开 {expanded_count} 个折叠的帖子")
            weibo_human_pause(0.8, 1.2)
    except Exception:
        pass


def weibo_is_repost_post(post_element) -> bool:
    try:
        repost_indicators = post_element.find_elements(
            By.CSS_SELECTOR,
            '[class*="repost"], [class*="forward"], [node-type="feed_list_forwardContent"], .card-comment',
        )
        for indicator in repost_indicators:
            indicator_text = indicator.text.strip().lower()
            if any(kw in indicator_text for kw in ["转发", "//@", "转发了"]):
                return True
        main_text = ""
        text_nodes = post_element.find_elements(By.CSS_SELECTOR, "p.txt")
        if text_nodes:
            main_text = text_nodes[0].text.strip()
        else:
            text_nodes = post_element.find_elements(By.CSS_SELECTOR, "p, div[dir='auto'], span")
            if text_nodes:
                main_text = text_nodes[0].text.strip()
            else:
                main_text = post_element.text.strip()
        if main_text.startswith("转发") or main_text.startswith("//@"):
            return True
        if re.search(r"//@\w+", main_text):
            return True
        repost_patterns = [r"^转发\s*微博", r"^转发了", r"转发\s*自", r"//@\w+"]
        for pattern in repost_patterns:
            if re.search(pattern, main_text, re.IGNORECASE):
                return True
        full_text = post_element.text.strip().lower()
        if any(keyword in full_text for keyword in ["转发微博", "//@", "转发了"]):
            if re.search(r"^(转发|//@)", full_text, re.MULTILINE):
                return True
        return False
    except Exception:
        return False


def weibo_collect_posts_from_dom(driver: webdriver.Chrome, expand_posts: bool = True) -> List[Dict[str, str]]:
    if expand_posts:
        try:
            weibo_expand_folded_posts(driver)
        except Exception:
            pass
    results: List[Dict[str, str]] = []
    try:
        cards = driver.find_elements(By.CSS_SELECTOR, "div.card-wrap")
        for c in cards:
            try:
                has_card = c.find_elements(By.CSS_SELECTOR, "div.card")
                if not has_card:
                    continue
                if weibo_is_repost_post(c):
                    continue
                content = ""
                text_nodes = c.find_elements(By.CSS_SELECTOR, "p.txt")
                if text_nodes:
                    main_text = text_nodes[0].text.strip()
                    main_text = weibo_filter_post_content(main_text)
                    content = weibo_clean_text(main_text)
                if not content or "展开" in content or len(content) < 10:
                    try:
                        text_containers = c.find_elements(
                            By.CSS_SELECTOR,
                            "p.txt, p[class*='txt'], div[class*='text'], div[dir='auto']",
                        )
                        text_parts = []
                        for elem in text_containers:
                            try:
                                parent = elem.find_element(By.XPATH, "./..")
                                parent_class = parent.get_attribute("class") or ""
                                if any(
                                    keyword in parent_class.lower()
                                    for keyword in ["image", "img", "photo", "pic", "vote", "poll", "app", "card"]
                                ):
                                    continue
                                txt = elem.text.strip()
                                if txt and len(txt) > 2:
                                    filtered = weibo_filter_post_content(txt)
                                    if filtered and len(filtered) > 2:
                                        text_parts.append(filtered)
                            except Exception:
                                continue
                        if text_parts:
                            unique_parts = []
                            seen = set()
                            for part in text_parts:
                                if part not in seen and len(part) > 3:
                                    seen.add(part)
                                    unique_parts.append(part)
                            if unique_parts:
                                full_content = " ".join(unique_parts)
                                if len(full_content) > len(content):
                                    content = weibo_clean_text(full_content)
                    except Exception:
                        pass
                interaction_data = weibo_extract_interaction_numbers(c)
                repost_text = interaction_data.get("Repost Count", "无数据")
                comment_text = interaction_data.get("Comment Count", "无数据")
                like_text = interaction_data.get("Like Count", "无数据")
                link = ""
                link_candidates = c.find_elements(By.CSS_SELECTOR, 'a[node-type="feed_list_item_date"], a[target="_blank"]')
                for a in link_candidates:
                    href = a.get_attribute("href") or ""
                    if "weibo.com" in href and re.search(r"/\d+/", href):
                        link = href
                        break
                if not link and link_candidates:
                    link = link_candidates[0].get_attribute("href") or ""
                if content:
                    results.append(
                        {
                            "Text Content": content or "",
                            "Repost Count": repost_text if (repost_text and repost_text.strip()) else "0",
                            "Comment Count": comment_text if (comment_text and comment_text.strip()) else "0",
                            "Like Count": like_text if (like_text and like_text.strip()) else "0",
                            "Post Link": link or "无数据",
                        }
                    )
            except Exception:
                continue
    except Exception:
        pass
    try:
        article_nodes = driver.find_elements(By.CSS_SELECTOR, "article, div[role='article']")
        for a in article_nodes:
            try:
                if weibo_is_repost_post(a):
                    continue
                content = ""
                text_containers = a.find_elements(
                    By.CSS_SELECTOR, "div[dir='auto'], p, span[class*='text'], div[class*='text']"
                )
                text_parts = []
                for elem in text_containers:
                    try:
                        try:
                            parent = elem.find_element(By.XPATH, "./..")
                            parent_class = (parent.get_attribute("class") or "").lower()
                            if any(
                                keyword in parent_class
                                for keyword in ["image", "img", "photo", "pic", "vote", "poll", "app", "card", "media"]
                            ):
                                continue
                        except Exception:
                            pass
                        txt = elem.text.strip()
                        if txt and len(txt) > 3 and "展开" not in txt:
                            filtered = weibo_filter_post_content(txt)
                            if filtered and len(filtered) > 3:
                                text_parts.append(filtered)
                    except Exception:
                        continue
                if text_parts:
                    unique_parts = []
                    seen = set()
                    for part in text_parts:
                        if part not in seen and len(part) > 3:
                            seen.add(part)
                            unique_parts.append(part)
                    if unique_parts:
                        content = weibo_clean_text(" ".join(unique_parts))
                if not content or len(content) < 10:
                    try:
                        main_text_nodes = a.find_elements(By.CSS_SELECTOR, "article > div, article > p, [role='textbox']")
                        for node in main_text_nodes:
                            txt = node.text.strip()
                            if txt:
                                filtered = weibo_filter_post_content(txt)
                                if filtered and len(filtered) > len(content):
                                    content = weibo_clean_text(filtered)
                    except Exception:
                        pass
                interaction_data = weibo_extract_interaction_numbers(a)
                repost_text = interaction_data.get("Repost Count", "无数据")
                comment_text = interaction_data.get("Comment Count", "无数据")
                like_text = interaction_data.get("Like Count", "无数据")
                link = ""
                link_candidates = a.find_elements(By.CSS_SELECTOR, "a")
                for l in link_candidates:
                    href = (l.get_attribute("href") or "").strip()
                    if "weibo.com" in href and re.search(r"/\d+", href):
                        link = href
                        break
                if content:
                    results.append(
                        {
                            "Text Content": content or "",
                            "Repost Count": repost_text if (repost_text and repost_text != "无数据" and repost_text.strip()) else "无数据",
                            "Comment Count": comment_text if (comment_text and comment_text != "无数据" and comment_text.strip()) else "无数据",
                            "Like Count": like_text if (like_text and like_text != "无数据" and like_text.strip()) else "无数据",
                            "Post Link": link or "无数据",
                        }
                    )
            except Exception:
                continue
    except Exception:
        pass
    return results


def weibo_scroll_or_next_page(driver: webdriver.Chrome) -> bool:
    try:
        next_btns = driver.find_elements(By.CSS_SELECTOR, "a.next, a.next_page, a[title='下一页']")
        for nb in next_btns:
            if nb.is_displayed() and nb.is_enabled():
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", nb)
                    weibo_human_pause(0.6, 1.2)
                    nb.click()
                    weibo_human_pause(*WEIBO_SCROLL_PAUSE_RANGE)
                    return True
                except Exception:
                    continue
    except Exception:
        pass
    try:
        last_height = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(weibo_rand_in_range(WEIBO_SCROLL_PAUSE_RANGE))
        new_height = driver.execute_script("return document.body.scrollHeight")
        return new_height > last_height
    except Exception:
        return False


def weibo_crawl_keyword(driver: webdriver.Chrome, keyword: str) -> Optional[List[Dict[str, str]]]:
    try:
        print(f"\n{'='*60}")
        print(f"开始爬取关键词：{keyword}")
        print(f"{'='*60}\n")
        try:
            main_window = driver.current_window_handle
            all_windows = driver.window_handles
            for window in all_windows:
                if window != main_window:
                    driver.switch_to.window(window)
                    driver.close()
            driver.switch_to.window(main_window)
        except Exception:
            pass
        print(f"正在跳转到搜索结果页面...")
        ok = weibo_try_homepage_search(driver, keyword)
        if not ok or not weibo_is_result_page(driver):
            print(f"首页搜索失败，使用直接跳转方式...")
            weibo_goto_search_result_direct(driver, keyword)
        try:
            main_window = driver.current_window_handle
            all_windows = driver.window_handles
            for window in all_windows:
                if window != main_window:
                    driver.switch_to.window(window)
                    driver.close()
            driver.switch_to.window(main_window)
        except Exception:
            pass
        current_url = driver.current_url
        print(f"当前页面URL: {current_url}")
        if not weibo_is_result_page(driver):
            print(f"警告：可能未成功到达搜索结果页面")
        weibo_human_pause(2.0, 3.0)
        collected: List[Dict[str, str]] = []
        seen_links = set()
        fail_streak = 0
        round_count = 0
        while len(collected) < WEIBO_TARGET_COUNT and fail_streak < WEIBO_PAGE_MAX_FAIL:
            round_count += 1
            print(f"\n第 {round_count} 轮收集...")
            weibo_human_pause(*WEIBO_SCROLL_PAUSE_RANGE)
            try:
                WebDriverWait(driver, WEIBO_ELEMENT_WAIT_SECONDS).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            except Exception as e:
                print(f"等待页面加载超时: {str(e)}")
            try:
                batch = weibo_collect_posts_from_dom(driver, expand_posts=True)
                print(f"本轮从页面提取到 {len(batch)} 条帖子")
            except Exception as e:
                print(f"提取帖子时出错: {str(e)}")
                batch = []
            new_added = 0
            for item in batch:
                link = item.get("Post Link", "") or ""
                key = link if link and link != "无数据" else hash(item.get("Text Content", ""))
                if key in seen_links:
                    continue
                seen_links.add(key)
                collected.append(item)
                new_added += 1
                if len(collected) >= WEIBO_TARGET_COUNT:
                    break
            print(f"当前已收集：{len(collected)}/{WEIBO_TARGET_COUNT} 条（本轮新增：{new_added} 条）")
            if new_added == 0:
                fail_streak += 1
                print(f"连续失败次数：{fail_streak}/{WEIBO_PAGE_MAX_FAIL}")
            else:
                fail_streak = 0
            try:
                progressed = weibo_scroll_or_next_page(driver)
                if not progressed:
                    fail_streak += 1
                    print(f"无法加载更多内容，连续失败次数：{fail_streak}/{WEIBO_PAGE_MAX_FAIL}")
            except Exception as e:
                print(f"滚动/翻页时出错: {str(e)}")
                fail_streak += 1
        collected = collected[:WEIBO_TARGET_COUNT]
        if len(collected) == 0:
            print(f"\n警告：关键词 '{keyword}' 未收集到任何数据！")
            print("  1. 搜索结果页面未正确加载")
            print("  2. 页面结构发生变化，无法识别帖子")
            print("  3. 需要登录或遇到验证码")
            print("  4. 网络连接问题")
            return None
        for row in collected:
            row.setdefault("Text Content", "")
            row.setdefault("Repost Count", "无数据")
            row.setdefault("Comment Count", "无数据")
            row.setdefault("Like Count", "无数据")
            row.setdefault("Post Link", "无数据")
        print(f"\n关键词 '{keyword}' 爬取完成！共收集 {len(collected)} 条数据\n")
        return collected
    except Exception as e:
        print(f"\n爬取关键词 '{keyword}' 时发生异常：{str(e)}")
        import traceback
        traceback.print_exc()
        return None


def weibo_save_combined_excel(all_data: List[tuple], out_dir: str):
    try:
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"微博搜索结果_整合_{date_str}.xlsx"
        filepath = os.path.join(out_dir, filename)
        combined_rows = []
        valid_data = [(k, d) for k, d in all_data if d]
        for keyword, data_list in valid_data:
            for row in data_list:
                combined_row = {"Key Word": keyword}
                combined_row.update(row)
                combined_rows.append(combined_row)
        if not combined_rows:
            print("警告：没有数据可导出！")
            return False
        df = pd.DataFrame(
            combined_rows,
            columns=["Key Word", "Text Content", "Repost Count", "Comment Count", "Like Count", "Post Link"],
        )
        for col in ["Repost Count", "Comment Count", "Like Count"]:
            df[col] = df[col].apply(lambda x: str(x) if x is not None and x != "" else ("无数据" if x == "" else ""))
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]
            for row in range(2, len(df) + 2):
                for col_idx in [3, 4, 5]:
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell_value = df.iloc[row - 2, col_idx - 1]
                    cell_value_str = str(cell_value) if cell_value is not None and cell_value != "" else ""
                    cell.number_format = "@"
                    cell.value = cell_value_str
        total_rows = sum(len(data_list) for _, data_list in valid_data)
        print(f"\n{'='*60}")
        print(f"整合Excel文件已导出：{filepath}")
        print(f"共包含 {len(valid_data)} 个关键词，总计 {total_rows} 条数据")
        print(f"{'='*60}\n")
        return True
    except Exception as e:
        print(f"\n导出整合Excel文件时出错：{str(e)}")
        import traceback
        traceback.print_exc()
        return False


def weibo_save_single_excel(keyword: str, data: List[Dict], out_dir: str):
    """按单个关键词导出微博数据到Excel。"""
    if not data:
        return
    try:
        os.makedirs(out_dir, exist_ok=True)
        filename = f"weibo_{_sanitize_filename(keyword)}.xlsx"
        filepath = os.path.join(out_dir, filename)
        df = pd.DataFrame(
            data,
            columns=["Text Content", "Repost Count", "Comment Count", "Like Count", "Post Link"],
        )
        for col in ["Repost Count", "Comment Count", "Like Count"]:
            df[col] = df[col].apply(lambda x: str(x) if x is not None and x != "" else ("无数据" if x == "" else ""))
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]
            for row in range(2, len(df) + 2):
                for col_idx in [2, 3, 4]:
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell_value = df.iloc[row - 2, col_idx - 1]
                    cell_value_str = str(cell_value) if cell_value is not None and cell_value != "" else ""
                    cell.number_format = "@"
                    cell.value = cell_value_str
        print(f"[Weibo] 单关键词Excel已导出 -> {filepath}")
    except Exception as e:
        print(f"[Weibo] 导出关键词 '{keyword}' Excel 时出错：{e}")


@dataclass
class NoteRecord:
    title: str
    like_count: str
    post_url: str


class XiaoHongShuScraper:
    def __init__(
        self,
        keyword: str,
        limit: int = 200,
        headless: bool = True,
        timeout_sec: int = 25,
        delay_min: float = 1.2,
        delay_max: float = 2.4,
        output_path: Optional[str] = None,
        proxy: Optional[str] = None,
    ):
        self.keyword = keyword
        self.limit = limit
        self.headless = headless
        self.timeout_sec = timeout_sec
        self.delay_min = delay_min
        self.delay_max = delay_max
        self.output_path = output_path
        self.proxy = proxy

        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None

    def _sleep_a_bit(self):
        delay = random.uniform(self.delay_min, self.delay_max)
        time.sleep(delay)

    def _ensure_output_path(self):
        if not self.output_path:
            os.makedirs(XHS_OUTPUT_DIR, exist_ok=True)
            filename = f"xhs_{_sanitize_filename(self.keyword)}.xlsx"
            self.output_path = os.path.join(XHS_OUTPUT_DIR, filename)

    def _launch(self, for_setup: bool = False):
        pw = sync_playwright().start()
        self._pw = pw

        launch_args: Dict = {
            "headless": (False if for_setup else self.headless),
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-sandbox",
            ],
        }
        if self.proxy:
            launch_args["proxy"] = {"server": self.proxy}

        self.browser = pw.chromium.launch(**launch_args)

        context_args: Dict = {
            "viewport": {"width": 1280, "height": 900},
            "user_agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "java_script_enabled": True,
        }
        if os.path.exists(STORAGE_STATE_FILE):
            context_args["storage_state"] = STORAGE_STATE_FILE

        self.context = self.browser.new_context(**context_args)
        self.page = self.context.new_page()

        # Try to reduce automation fingerprints
        self.page.add_init_script(
            """
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            """
        )

    def _close(self, save_state: bool = False):
        """
        关闭浏览器和上下文
        
        参数:
            save_state: 是否保存登录状态（默认False，因为通常已在外部保存）
        """
        try:
            if save_state and self.context:
                try:
                    self.context.storage_state(path=STORAGE_STATE_FILE)
                    if os.path.exists(STORAGE_STATE_FILE):
                        file_size = os.path.getsize(STORAGE_STATE_FILE)
                        console.print(f"[info] 已保存登录状态到 {STORAGE_STATE_FILE} (大小: {file_size} 字节)")
                    else:
                        console.print(f"[warn] 保存登录状态到 {STORAGE_STATE_FILE}，但文件未创建")
                except Exception as e:
                    console.print(f"[warn] 保存登录状态时出错: {e}")
                    console.print(f"[debug] 错误详情: {traceback.format_exc()}")
        finally:
            try:
                if self.context:
                    self.context.close()
                if self.browser:
                    self.browser.close()
            finally:
                if hasattr(self, "_pw"):
                    self._pw.stop()

    def _check_login_status(self) -> bool:
        """
        检查是否已登录（检查cookie和页面状态）
        
        返回:
            bool: 如果检测到有效的登录状态，返回True
        """
        try:
            if not self.context or not self.page:
                return False
            
            # 方法1: 检查cookie - 小红书的关键登录cookie
            cookies = self.context.cookies()
            cookie_dict = {c.get('name', '').lower(): c.get('value', '') for c in cookies}
            
            # 小红书的关键登录cookie
            login_indicators = [
                'web_session', 'web_session_2', 'sessionid', 
                'a1', 'webid', 'webid_2', 'websectiga',
                'xsecappid', 'sec_poison_id'
            ]
            
            found_indicators = 0
            for indicator in login_indicators:
                if indicator in cookie_dict and cookie_dict[indicator]:
                    found_indicators += 1
            
            # 至少需要1个关键cookie才认为可能已登录
            if found_indicators == 0:
                return False
            
            # 方法2: 检查页面元素 - 查找登录后的特征元素
            try:
                # 等待页面加载
                self.page.wait_for_load_state("domcontentloaded", timeout=5000)
                
                # 检查是否有用户头像、用户名等登录后的元素
                login_elements = [
                    "[class*='avatar']",
                    "[class*='user']",
                    "[class*='nickname']",
                    "img[alt*='头像']",
                    "[data-v-] img[src*='avatar']"
                ]
                
                for selector in login_elements:
                    try:
                        element = self.page.query_selector(selector)
                        if element and element.is_visible():
                            return True
                    except Exception:
                        continue
                
                # 检查是否有登录按钮（如果有说明未登录）
                login_buttons = [
                    "button:has-text('登录')",
                    "a:has-text('登录')",
                    "button:has-text('立即登录')",
                    "[class*='login'] button"
                ]
                
                for selector in login_buttons:
                    try:
                        element = self.page.query_selector(selector)
                        if element and element.is_visible():
                            return False
                    except Exception:
                        continue
            except Exception:
                pass
            
            # 方法3: 检查页面URL，如果被重定向到登录页，说明未登录
            try:
                current_url = self.page.url.lower()
                if "login" in current_url or "signin" in current_url or "/passport" in current_url:
                    return False
            except Exception:
                pass
            
            # 方法4: 检查页面标题或内容中是否有登录提示
            try:
                page_title = self.page.title().lower()
                login_keywords = ['登录', 'login', 'sign in', '请登录', 'signin']
                if any(keyword in page_title for keyword in login_keywords):
                    return False
            except Exception:
                pass
            
            # 如果找到了登录相关的cookie，且没有明显的未登录标志，认为已登录
            return True
        except Exception as e:
            console.print(f"[debug] 检查登录状态时出错: {e}")
            return False

    def run_setup(self):
        """
        预登录功能：打开浏览器让用户手动登录，然后保存登录状态
        """
        console.print("\n" + "=" * 60)
        console.print("[setup] 开始预登录流程...")
        console.print("=" * 60)
        console.print("[setup] 正在打开浏览器...")
        
        self._launch(for_setup=True)
        assert self.page
        
        # 访问小红书首页
        console.print("[setup] 正在访问小红书首页...")
        try:
            self.page.goto("https://www.xiaohongshu.com", timeout=self.timeout_sec * 1000)
            self.page.wait_for_load_state("domcontentloaded", timeout=self.timeout_sec * 1000)
            time.sleep(2)  # 等待页面完全加载
        except Exception as e:
            console.print(f"[warn] 访问首页时出错: {e}，继续...")
        
        console.print("\n" + "=" * 60)
        console.print("[setup] 请在打开的浏览器中手动登录小红书账号")
        console.print("[setup] 登录方式：扫码登录或账号密码登录")
        console.print("[setup] 登录完成后，脚本将自动检测并保存登录状态")
        console.print("=" * 60)
        console.print("[setup] 等待登录中...（最多等待180秒，可以按 Ctrl+C 提前结束）")
        
        wait_time = 180  # 等待180秒
        elapsed = 0
        login_detected = False
        consecutive_checks = 0  # 连续检测到登录状态的次数
        
        try:
            while elapsed < wait_time:
                time.sleep(3)  # 每3秒检查一次
                elapsed += 3
                
                # 每10秒提示一次
                if elapsed % 10 == 0:
                    remaining = wait_time - elapsed
                    console.print(f"[setup] 已等待 {elapsed} 秒，剩余 {remaining} 秒...")
                
                # 检查是否已登录
                if self._check_login_status():
                    consecutive_checks += 1
                    # 连续2次检测到登录状态才确认（避免误判）
                    if consecutive_checks >= 2:
                        login_detected = True
                        console.print("[ok] ✓ 检测到登录状态！")
                        # 等待几秒确保登录流程完全完成
                        console.print("[info] 等待登录流程完全完成...")
                        time.sleep(3)
                        
                        # 再次访问首页确保状态稳定
                        try:
                            self.page.goto("https://www.xiaohongshu.com", timeout=self.timeout_sec * 1000)
                            self.page.wait_for_load_state("domcontentloaded", timeout=self.timeout_sec * 1000)
                            time.sleep(2)
                        except Exception:
                            pass
                        break
                else:
                    consecutive_checks = 0  # 重置计数
                    
        except KeyboardInterrupt:
            console.print("\n[info] 用户中断，正在检查当前状态...")
            # 最后再检查一次
            if self._check_login_status():
                login_detected = True
                console.print("[ok] 检测到登录状态")
        
        # 如果检测到登录状态，保存登录状态
        if login_detected:
            console.print("\n" + "=" * 60)
            console.print("[info] 开始保存登录状态...")
            console.print("=" * 60)
            
            try:
                if self.context:
                    # 确保页面已加载
                    try:
                        self.page.wait_for_load_state("networkidle", timeout=10000)
                    except Exception:
                        pass
                    
                    abs_path = os.path.abspath(STORAGE_STATE_FILE)
                    console.print(f"[info] 正在保存登录状态到: {abs_path}")
                    
                    # 保存登录状态
                    self.context.storage_state(path=STORAGE_STATE_FILE)
                    
                    # 等待文件写入完成
                    time.sleep(1)
                    
                    # 验证文件是否成功创建
                    if os.path.exists(STORAGE_STATE_FILE):
                        file_size = os.path.getsize(STORAGE_STATE_FILE)
                        if file_size > 100:  # 文件大小应该至少100字节
                            console.print(f"[ok] ✓ 登录状态已成功保存！")
                            console.print(f"[info] 文件路径: {abs_path}")
                            console.print(f"[info] 文件大小: {file_size} 字节")
                            console.print("[info] 下次运行时将自动使用保存的登录状态。")
                        else:
                            console.print(f"[warn] 文件大小异常 ({file_size} 字节)，保存可能失败")
                    else:
                        console.print("[error] 文件未创建，保存失败")
                else:
                    console.print("[error] Context对象不存在，无法保存")
            except Exception as e:
                console.print(f"[error] 保存登录状态失败: {e}")
                console.print(f"[debug] 错误详情: {traceback.format_exc()}")
        else:
            console.print("\n" + "=" * 60)
            console.print("[warn] 未检测到有效的登录状态")
            console.print("[info] 可能的原因：")
            console.print("  1. 未完成登录流程")
            console.print("  2. 登录状态检测失败")
            console.print("  3. 登录已过期")
            console.print("=" * 60)
            console.print("[info] 请重新运行 --setup 完成登录。")
        
        # 关闭浏览器
        console.print("\n[info] 正在关闭浏览器...")
        self._close(save_state=False)
        console.print("[ok] 浏览器已关闭")
        console.print("=" * 60)

    def _go_search(self):
        assert self.page
        search_url = SEARCH_URL_TEMPLATE.format(keyword=self.keyword)
        self.page.goto(search_url, timeout=self.timeout_sec * 1000)
        self.page.wait_for_load_state("domcontentloaded", timeout=self.timeout_sec * 1000)
        self._sleep_a_bit()

        # 若有显式搜索框也尝试输入一下（容错）
        try:
            search_input = self.page.query_selector("input[placeholder*='搜索']")
            if search_input:
                search_input.fill("")
                search_input.type(self.keyword, delay=50)
                self._sleep_a_bit()
                search_input.press("Enter")
                self.page.wait_for_load_state("networkidle", timeout=self.timeout_sec * 1000)
        except Exception:
            pass

        # 尝试切换到"笔记"标签，确保仅抓取搜索结果中的帖子
        try:
            # 兼容按钮/标签两种写法
            for sel in [
                "button:has-text('笔记')",
                "[role='tab']:has-text('笔记')",
                "a:has-text('笔记')",
            ]:
                btn = self.page.query_selector(sel)
                if btn:
                    btn.click()
                    self.page.wait_for_load_state("networkidle", timeout=self.timeout_sec * 1000)
                    self._sleep_a_bit()
                    break
        except Exception:
            pass

        # 尝试按点赞数排序
        console.print("[info] 尝试按点赞数排序...")
        try:
            # 尝试找到排序按钮/下拉框
            sort_selectors = [
                "button:has-text('最热')",
                "button:has-text('点赞')",
                "[class*='sort'] button",
                "[class*='filter'] button",
                "button:has-text('排序')",
            ]
            for sel in sort_selectors:
                btn = self.page.query_selector(sel)
                if btn:
                    btn.click()
                    self._sleep_a_bit()
                    # 如果有下拉菜单，选择"按点赞数"
                    like_sort_options = [
                        "button:has-text('点赞')",
                        "li:has-text('点赞')",
                        "a:has-text('点赞')",
                        "[class*='option']:has-text('点赞')",
                    ]
                    for opt_sel in like_sort_options:
                        opt = self.page.query_selector(opt_sel)
                        if opt:
                            opt.click()
                            self.page.wait_for_load_state("networkidle", timeout=self.timeout_sec * 1000)
                            self._sleep_a_bit()
                            console.print("[ok] 已切换到按点赞数排序")
                            return
                    break
        except Exception as e:
            console.print(f"[warn] 无法切换到按点赞数排序: {e}，将按默认排序继续")
            pass

    def _extract_note_from_card(self, card_element) -> Optional[NoteRecord]:
        """从搜索结果页的卡片元素中提取标题、点赞数和链接"""
        try:
            # 提取链接
            link_elem = card_element.query_selector("a[href^='/explore/']")
            if not link_elem:
                return None
            
            href = link_elem.get_attribute("href") or ""
            if not href.startswith("/explore/"):
                return None
            
            post_url = NOTE_URL_PREFIX + href.split("/explore/")[-1]
            
            # 提取标题 - 可能在卡片内多个位置
            title = ""
            title_selectors = [
                "[class*='title']",
                "[class*='desc']",
                "h3",
                "h2",
                "p",
            ]
            for sel in title_selectors:
                title_elem = card_element.query_selector(sel)
                if title_elem:
                    title = title_elem.inner_text().strip()
                    if title and len(title) > 2:
                        break
            
            # 提取点赞数
            like_count = ""
            like_selectors = [
                "[class*='like']",
                "[class*='count']",
                "span:has-text('赞')",
                "button:has(svg) span",
            ]
            for sel in like_selectors:
                like_elems = card_element.query_selector_all(sel)
                for like_elem in like_elems:
                    text = like_elem.inner_text().strip()
                    # 尝试从文本中提取数字
                    match = re.search(r'(\d+(?:\.\d+)?[万万千]?)', text.replace(",", ""))
                    if match:
                        like_count = match.group(1)
                        break
                if like_count:
                    break
            
            # 如果没找到点赞数，尝试在整个卡片中搜索数字
            if not like_count:
                card_text = card_element.inner_text()
                matches = re.findall(r'(\d+(?:\.\d+)?[万万千]?)\s*[赞赞]', card_text)
                if matches:
                    like_count = matches[0]
            
            if title:
                return NoteRecord(
                    title=title,
                    like_count=like_count or "0",
                    post_url=post_url,
                )
        except Exception as e:
            pass
        return None

    def _scroll_and_collect_notes(self) -> List[NoteRecord]:
        """在搜索结果页滚动并收集笔记信息"""
        assert self.page
        seen: set = set()
        records: List[NoteRecord] = []

        last_height = 0
        stalled_rounds = 0
        scroll_count = 0
        
        console.print(f"[info] 开始滚动页面并收集笔记，目标：{self.limit} 条")
        
        while len(records) < self.limit and stalled_rounds < 8:
            scroll_count += 1
            prev_count = len(records)
            
            # 查找所有帖子卡片容器
            card_selectors = [
                "[class*='note-item']",
                "[class*='feed-item']",
                "[class*='card']",
                "article",
                "a[href^='/explore/']",
            ]
            
            cards_found = False
            for card_sel in card_selectors:
                cards = self.page.query_selector_all(card_sel)
                if cards:
                    cards_found = True
                    for card in cards:
                        try:
                            # 检查可见性
                            if not card.is_visible():
                                continue
                        except Exception:
                            pass
                        
                        # 提取笔记信息
                        note = self._extract_note_from_card(card)
                        if note and note.post_url not in seen:
                            seen.add(note.post_url)
                            records.append(note)
                            # 实时显示采集进度
                            console.print(f"[progress] 已收集 {len(records)}/{self.limit} 条 | 标题: {note.title[:30]}... | 点赞: {note.like_count}")
                            if len(records) >= self.limit:
                                break
                    if len(records) >= self.limit:
                        break
                if len(records) >= self.limit:
                    break
            
            if len(records) >= self.limit:
                console.print(f"[ok] 已达到目标数量 {self.limit} 条")
                break

            # 向下滚动加载
            console.print(f"[info] 第 {scroll_count} 次滚动，当前已收集 {len(records)} 条...")
            self.page.evaluate("window.scrollBy(0, document.body.scrollHeight * 0.5)")
            self._sleep_a_bit()
            self.page.wait_for_timeout(800)

            new_height = self.page.evaluate("document.body.scrollHeight")
            if new_height == last_height:
                stalled_rounds += 1
                console.print(f"[warn] 页面高度未变化，停滞轮次: {stalled_rounds}/8")
            else:
                stalled_rounds = 0
                last_height = new_height
                if len(records) > prev_count:
                    console.print(f"[info] 本轮新增 {len(records) - prev_count} 条笔记")

        console.print(f"[ok] 收集完成，共收集笔记：{len(records)} 条")
        return records[: self.limit]



    def scrape(self) -> List[NoteRecord]:
        self._ensure_output_path()
        console.print(f"[info] 启动浏览器{'（无头模式）' if self.headless else '（可见模式，可在浏览器窗口中观察）'}...")
        
        # 检查是否有保存的登录状态
        if os.path.exists(STORAGE_STATE_FILE):
            file_size = os.path.getsize(STORAGE_STATE_FILE)
            console.print(f"[info] 检测到已保存的登录状态（{file_size} 字节），正在加载...")
        else:
            console.print("[warn] 未检测到保存的登录状态，建议先运行 --setup 进行预登录")
        
        self._launch(for_setup=False)
        assert self.page
        
        try:
            # 访问首页验证登录状态
            try:
                console.print("[info] 正在验证登录状态...")
                self.page.goto("https://www.xiaohongshu.com", timeout=self.timeout_sec * 1000)
                self.page.wait_for_load_state("domcontentloaded", timeout=self.timeout_sec * 1000)
                time.sleep(1)
            except Exception as e:
                console.print(f"[warn] 访问首页时出错: {e}")
            
            # 检查登录状态
            if self._check_login_status():
                console.print("[ok] ✓ 登录状态正常")
            else:
                console.print("[warn] ⚠ 未检测到登录状态，某些功能可能受限")
                console.print("[info] 提示：运行 'python 自动化.py --setup' 进行预登录")
            
            console.print(f"[info] 正在搜索关键词: {self.keyword}")
            self._go_search()
            console.print(f"[info] 开始收集笔记，目标数量：{self.limit}...")
            records = self._scroll_and_collect_notes()
            return records
        finally:
            self._close()

    def export_excel(self, records: List[NoteRecord]):
        if not records:
            console.print("[warn] 无任何记录可导出。")
            return
        
        # 转换为字典列表
        data = [asdict(r) for r in records]
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 确保输出目录存在
        output_dir = os.path.dirname(self.output_path or XHS_OUTPUT_DIR)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        
        # 导出为Excel
        df.to_excel(self.output_path, index=False, engine='openpyxl')
        console.print(f"[ok] Excel 导出完成 -> {self.output_path}")


def xhs_save_combined_excel(all_data: List[tuple], out_dir: str):
    """小红书多关键词整合导出。"""
    try:
        os.makedirs(out_dir, exist_ok=True)
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"xhs_整合_{date_str}.xlsx"
        filepath = os.path.join(out_dir, filename)

        combined_rows = []
        for keyword, records in all_data:
            if not records:
                continue
            for r in records:
                row = asdict(r)
                row["Keyword"] = keyword
                combined_rows.append(row)

        if not combined_rows:
            print("[XHS] 无数据可导出整合文件。")
            return False

        df = pd.DataFrame(combined_rows, columns=["Keyword", "title", "like_count", "post_url"])
        df.rename(columns={"title": "Title", "like_count": "Like Count", "post_url": "Post URL"}, inplace=True)
        df.to_excel(filepath, index=False, engine="openpyxl")
        print(f"[XHS] 整合Excel已导出 -> {filepath}")
        return True
    except Exception as e:
        print(f"[XHS] 导出整合Excel时出错：{e}")
        return False


class WeiboRunner:
    """
    内联的微博爬虫，按关键词批量爬取微博搜索结果并导出。
    """

    def __init__(self, limit: int = 200, headless: bool = False):
        self.limit = limit
        self.headless = headless
        self.driver = None
        self.all_keyword_data = []

        # 配置内联微博参数
        global WEIBO_TARGET_COUNT, WEIBO_HEADLESS
        WEIBO_TARGET_COUNT = self.limit
        WEIBO_HEADLESS = self.headless

    def _init_driver(self):
        self.driver = weibo_build_driver()
        weibo_inject_manual_cookie_if_any(self.driver)

        # 登录检测（仅在可视化模式下提示扫码）
        self.driver.get("https://www.weibo.com/")
        try:
            WebDriverWait(self.driver, WEIBO_ELEMENT_WAIT_SECONDS).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
        except Exception:
            pass

        if not WEIBO_MANUAL_COOKIE_STRING and not self.headless:
            if not weibo_is_logged_in(self.driver):
                print(
                    f"请在弹出的浏览器中扫码登录微博，系统将等待最多 {WEIBO_QR_LOGIN_WAIT_SECONDS} 秒..."
                )
                ok_login = weibo_wait_for_manual_login(
                    self.driver, WEIBO_QR_LOGIN_WAIT_SECONDS
                )
                try:
                    self.driver.get("https://www.weibo.com/")
                    time.sleep(1.0)
                except Exception:
                    pass
                if not ok_login and not weibo_is_logged_in(self.driver):
                    print("未检测到微博登录，微博爬虫将跳过。")
                    return False
        return True

    def run_keywords(self, keywords: List[str]) -> int:
        """
        批量爬取关键词，返回成功的关键词数量。
        """
        if not keywords:
            print("微博关键词列表为空，跳过微博爬取。")
            return 0

        if not self.driver:
            ok = self._init_driver()
            if not ok:
                return 0

        success_count = 0
        for idx, keyword in enumerate(keywords, 1):
            print(f"\n[Weibo] 进度：{idx}/{len(keywords)} | 关键词：{keyword}")
            data = weibo_crawl_keyword(self.driver, keyword)
            if data:
                self.all_keyword_data.append((keyword, data))
                weibo_save_single_excel(keyword, data, WEIBO_OUTPUT_DIR)
                success_count += 1

            # 关键词之间稍作停顿
            if idx < len(keywords):
                pause = weibo_rand_in_range((2.0, 4.0))
                print(f"[Weibo] 等待 {pause:.1f} 秒后处理下一个关键词...")
                weibo_human_pause(2.0, 4.0)

        # 导出整合文件
        if self.all_keyword_data:
            os.makedirs(WEIBO_OUTPUT_DIR, exist_ok=True)
            out_dir = os.path.abspath(WEIBO_OUTPUT_DIR)
            weibo_save_combined_excel(self.all_keyword_data, out_dir)

        return success_count

    def close(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="自动化爬虫工具：中关村爬虫 + 小红书/微博爬虫")
    
    # 自动化工作流参数
    parser.add_argument("--auto", action="store_true", help="运行自动化工作流（step1中关村爬虫 -> step2小红书/微博爬虫）")
    parser.add_argument("--excel-file", type=str, default=None, help="指定中关村爬虫输出的Excel文件路径（用于--auto模式）")
    parser.add_argument("--skip-step1", action="store_true", help="跳过step1，直接使用现有Excel文件运行step2（需要配合--excel-file使用）")
    parser.add_argument("--zol-start-page", type=int, default=1, help="中关村爬虫起始页码（用于--auto模式）")
    parser.add_argument("--zol-end-page", type=int, default=3, help="中关村爬虫结束页码（用于--auto模式）")
    
    # 小红书爬虫参数（用于--auto模式）
    parser.add_argument("--xhs-limit", type=int, default=200, help="小红书爬虫每个关键词的采集数量上限（用于--auto模式）")
    parser.add_argument("--xhs-headless", dest="xhs_headless", action="store_true", help="小红书爬虫使用无头模式（用于--auto模式）")
    parser.add_argument("--xhs-no-headless", dest="xhs_headless", action="store_false", help="小红书爬虫使用有头模式（用于--auto模式）")
    parser.add_argument("--no-xhs", dest="run_xhs", action="store_false", help="跳过小红书爬虫（用于--auto模式）")

    # 微博爬虫参数（用于--auto模式）
    parser.add_argument("--weibo-limit", type=int, default=200, help="微博爬虫每个关键词的采集数量上限（用于--auto模式）")
    parser.add_argument("--weibo-headless", dest="weibo_headless", action="store_true", help="微博爬虫使用无头模式（用于--auto模式）")
    parser.add_argument("--weibo-no-headless", dest="weibo_headless", action="store_false", help="微博爬虫使用有头模式（用于--auto模式）")
    parser.add_argument("--no-weibo", dest="run_weibo", action="store_false", help="跳过微博爬虫（用于--auto模式）")
    
    # 豆瓣爬虫参数（用于--auto模式）
    parser.add_argument("--douban-headless", dest="douban_headless", action="store_true", help="豆瓣爬虫使用无头模式（用于--auto模式）")
    parser.add_argument("--douban-no-headless", dest="douban_headless", action="store_false", help="豆瓣爬虫使用有头模式（用于--auto模式）")
    parser.add_argument("--no-douban", dest="run_douban", action="store_false", help="跳过豆瓣爬虫（用于--auto模式）")
    
    # 知乎爬虫参数（用于--auto模式）
    parser.add_argument("--zhihu-headless", dest="zhihu_headless", action="store_true", help="知乎爬虫使用无头模式（用于--auto模式）")
    parser.add_argument("--zhihu-no-headless", dest="zhihu_headless", action="store_false", help="知乎爬虫使用有头模式（用于--auto模式）")
    parser.add_argument("--no-zhihu", dest="run_zhihu", action="store_false", help="跳过知乎爬虫（用于--auto模式）")
    
    parser.set_defaults(
        xhs_headless=True,
        run_xhs=True,
        weibo_headless=False,
        run_weibo=True,
        douban_headless=False,
        run_douban=True,
        zhihu_headless=False,
        run_zhihu=True,
    )
    
    # 单独运行小红书爬虫的参数
    parser.add_argument("--keyword", type=str, default=None, help="搜索关键词（单独运行小红书爬虫时使用）")
    parser.add_argument("--limit", type=int, default=200, help="采集数量上限（单独运行小红书爬虫时使用）")
    parser.add_argument("--headless", dest="headless", action="store_true", help="无头模式（后台运行，不可见浏览器）")
    parser.add_argument("--no-headless", dest="headless", action="store_false", help="有头模式（显示浏览器窗口，可观察采集过程）")
    parser.set_defaults(headless=True)
    parser.add_argument("--timeout", type=int, default=25, help="单页等待超时(秒)")
    parser.add_argument("--delay-min", type=float, default=1.2, help="最小延时(秒)")
    parser.add_argument("--delay-max", type=float, default=2.4, help="最大延时(秒)")
    parser.add_argument("--output", type=str, default=None, help="Excel 输出路径")
    parser.add_argument("--proxy", type=str, default=None, help="HTTP/SOCKS 代理，例如 http://host:port 或 socks5://host:port")
    parser.add_argument("--setup", action="store_true", help="预登录模式：打开浏览器进行登录并保存登录状态")
    parser.add_argument("--douban-setup", action="store_true", help="豆瓣预登录：仅登录豆瓣并保存cookies，供后续爬虫复用")
    
    return parser.parse_args()


def load_keyword_from_env() -> Optional[str]:
    load_dotenv()
    kw = os.getenv("KEYWORD")
    if kw and kw.strip():
        return kw.strip()
    return None


def main_xhs():
    """单独运行小红书爬虫的主函数"""
    args = parse_args()
    
    # 如果是setup模式，不需要关键词，直接运行登录设置
    if args.setup:
        scraper = XiaoHongShuScraper(
            keyword="setup",  # 临时关键词，不会实际使用
            limit=args.limit,
            headless=False,  # setup模式必须显示浏览器
            timeout_sec=args.timeout,
            delay_min=args.delay_min,
            delay_max=args.delay_max,
            output_path=args.output,
            proxy=args.proxy,
        )
        scraper.run_setup()
        return
    
    # 非setup模式需要关键词
    keyword = args.keyword or load_keyword_from_env()
    if not keyword:
        console.print("[error] 未提供关键词。请使用 --keyword 或在 .env 中设置 KEYWORD。")
        sys.exit(1)

    scraper = XiaoHongShuScraper(
        keyword=keyword,
        limit=args.limit,
        headless=args.headless,
        timeout_sec=args.timeout,
        delay_min=args.delay_min,
        delay_max=args.delay_max,
        output_path=args.output,
        proxy=args.proxy,
    )

    records = scraper.scrape()
    scraper.export_excel(records)


def main():
    """主入口函数"""
    args = parse_args()
    
    # 豆瓣预登录模式：仅完成登录状态保存
    if args.douban_setup:
        douban_run_setup()
        return

    # 如果指定了--auto参数，运行自动化工作流
    if args.auto:
        run_automated_workflow(
            zol_start_page=args.zol_start_page,
            zol_end_page=args.zol_end_page,
            excel_file_path=args.excel_file,
            xhs_limit=args.xhs_limit,
            xhs_headless=args.xhs_headless,
            run_xhs=args.run_xhs,
            skip_step1=args.skip_step1,
            run_weibo=args.run_weibo,
            weibo_limit=args.weibo_limit,
            weibo_headless=args.weibo_headless,
            run_douban=args.run_douban,
            douban_headless=args.douban_headless,
            run_zhihu=args.run_zhihu,
            zhihu_headless=args.zhihu_headless,
        )
    else:
        # 否则运行小红书爬虫
        main_xhs()


if __name__ == "__main__":
    main()
